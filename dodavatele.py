#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
dodavatele.py - obohaceni seznamu dodavatelu z verejnych rejstriku.

Vstup : CSV / XLSX / TXT se seznamem nazvu firem (volitelne ICO, DIC, zeme).
Vystup: XLSX / CSV se sloupci
        Jmeno | Ulice | PSC | Mesto | Zeme | ICO | DIC | NACE
        | Kategorie dodavatele | ...

Zdroje (vse bez API klice):
  ARES     ares.gov.cz            CR - plny rejstrik vc. NACE a DIC
  RPO SR   api.statistics.sk      SR - registr pravnickych osob
  GLEIF    api.gleif.org          svet - entity s LEI, narodni registracni cislo
Handelsregister offeneregister.de  DE - lokalni kopie (otevrena data, k 2019)
  SEC      sec.gov                US - spolecnosti registrovane u SEC (+ SIC)
  Wikidata wikidata.org           zalozni zdroj oboru, DIC a sidla pro velke firmy
  VIES     ec.europa.eu           overeni DIC v ramci EU

Zavislosti: pouze standardni knihovna (openpyxl jen pro praci s XLSX).
"""

from __future__ import annotations

import argparse
import csv
import gzip
import io
import json
import os
import re
import sqlite3
import ssl
import sys
import threading
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from difflib import SequenceMatcher

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import taxonomie

UA = "supplier-lookup/1.0 (kontakt: nakup@example.com)"

ARES_HLEDAT = "https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty/vyhledat"
ARES_DETAIL = "https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty/{ico}"
ARES_RES = "https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty-res/{ico}"
ARES_VR = "https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty-vr/{ico}"
RPO_SK = "https://api.statistics.sk/rpo/v1/search"
GLEIF_API = "https://api.gleif.org/api/v1/lei-records"
GLEIF_AUTO = "https://api.gleif.org/api/v1/autocompletions"
VIES_API = "https://ec.europa.eu/taxation_customs/vies/rest-api/ms/{cc}/vat/{num}"
EDGAR_API = "https://www.sec.gov/cgi-bin/browse-edgar"
WIKIDATA_API = "https://www.wikidata.org/w/api.php"
INSEE_FR = "https://recherche-entreprises.api.gouv.fr/search"
SG_ACRA = "https://data.gov.sg/api/action/datastore_search"
SG_ACRA_ZDROJ = "d_3f960c10fed6145404ca7b821f263b87"
TW_GCIS = "https://data.gcis.nat.gov.tw/od/data/api/6BBA2268-1367-4B42-9CCA-BC17499EBE8C"

_TW_SSL_KONTEXT = None


def tw_ssl_kontext():
    """
    Certifikat data.gcis.nat.gov.tw postrada rozsireni Subject Key Identifier,
    ktere novejsi OpenSSL/Python defaultne vyzaduje (VERIFY_X509_STRICT).
    Overeni retezce duvery a jmena hostitele zustava aktivni - vypina se jen
    tato jedna nadstandardni RFC 5280 kontrola, ktera zpusobuje, ze pripojeni
    ze standardniho urllib kontextu vzdy selze, i kdyz je server v poradku
    (napr. curl tuto kontrolu vubec neprovadi, proto tam problem videt neni).
    """
    global _TW_SSL_KONTEXT
    if _TW_SSL_KONTEXT is None:
        ctx = ssl.create_default_context()
        ctx.verify_flags &= ~ssl.VERIFY_X509_STRICT
        _TW_SSL_KONTEXT = ctx
    return _TW_SSL_KONTEXT

EU_STATY = {"AT", "BE", "BG", "CY", "CZ", "DE", "DK", "EE", "EL", "GR", "ES", "FI", "FR",
            "HR", "HU", "IE", "IT", "LT", "LU", "LV", "MT", "NL", "PL", "PT", "RO", "SE",
            "SI", "SK", "XI"}

PRAVNI_FORMY = {
    "as", "a s", "akciova spolecnost", "sro", "s r o", "spol s r o",
    "spolecnost s rucenim omezenym", "ks", "vos", "zs", "ops", "sp", "se", "statni podnik",
    "zapsany spolek", "zu", "odstepny zavod",
    "gmbh", "mbh", "ag", "kg", "kgaa", "ohg", "ug", "eg", "ev", "gbr", "co kg", "se co kg",
    "aktiengesellschaft", "gesellschaft mit beschrankter haftung", "kommanditgesellschaft",
    "aktiebolag", "aktieselskab", "naamloze vennootschap", "besloten vennootschap",
    "societe anonyme", "societa per azioni", "sociedad anonima",
    "akciova spolocnost", "spolocnost s rucenim obmedzenym",
    "bv", "nv", "cv", "vof", "sa", "sas", "sasu", "sarl", "eurl", "sci",
    "srl", "spa", "snc", "sl", "slu", "sau",
    "ltd", "limited", "plc", "llp", "lp", "llc", "inc", "incorporated", "corp",
    "corporation", "company", "co", "holding", "holdings", "group", "groupe", "gruppe",
    "oy", "oyj", "ab", "aps", "asa", "kft", "zrt", "nyrt", "bt", "kkt",
    "sp z oo", "spzoo", "z oo", "doo", "dooel", "dd", "ood", "eood", "ad", "ead",
    "lda", "unipessoal", "kk", "yk", "pte", "pty", "bhd", "sdn", "akciova spolocnost",
    # viceslovne narodni formy - bez nich se "ORLEN S.A." neshodne
    # s "ORLEN SPOLKA AKCYJNA" ani "Vodafone Group Plc" s "VODAFONE GROUP
    # PUBLIC LIMITED COMPANY"
    "public limited company", "public company limited", "company limited", "co ltd",
    "joint stock company", "designated activity company", "dac",
    "spolka akcyjna", "spolka z ograniczona odpowiedzialnoscia", "sp z o o",
    "societa per azioni", "societa a responsabilita limitata",
    "sociedad anonima unipersonal", "sociedade anonima",
    "societe par actions simplifiee", "societe a responsabilite limitee",
    "anonim sirketi", "anonim ortakligi", "limited sirketi", "sirketi",
    "reszvenytarsasag", "nyilvanosan mukodo reszvenytarsasag",
    "societate pe actiuni", "societate cu raspundere limitata",
    "kabushiki kaisha", "berhad", "sendirian berhad",
    "naamloze vennootschap", "besloten vennootschap",
    "publikt aktiebolag", "aktiengesellschaft", "eingetragener verein",
}

STAV_OK = "OK"
STAV_VYBRANO = "VYBRANO"        # vice srovnatelnych shod, nastroj vybral nejlepsi
STAV_OVERIT = "OVERIT"
STAV_NENALEZENO = "NENALEZENO"
STAV_CHYBA = "CHYBA"

# ARES kody pravni formy pro fyzickou osobu podnikajici (OSVC) - viz RES
# ciselnik pravnich forem. Jmeno osoby (na rozdil od nazvu firmy) neni
# jednoznacny identifikator - u bezneho jmena (Jan Novak) ARES bezne eviduje
# stovky ruznych lidi (napr. "Jan Novak" ma v ARES 428 zaznamu) - proto se
# shoda podle pouheho jmena OSVC bez adresy/ICO k rozliseni nebere jako OK,
# viz vyber_nejlepsi().
OSVC_PRAVNI_FORMY = {"100", "101", "102", "103", "104", "105", "106", "107", "108"}


# ---------------------------------------------------------------------------
# HTTP vrstva: rate limit na host, opakovani pri chybe, cache na disku
# ---------------------------------------------------------------------------

class Klient:
    def __init__(self, cache_soubor=None, prodleva=0.25, pokusy=3, timeout=25, ua=UA):
        self.prodleva = prodleva
        self.pokusy = pokusy
        self.timeout = timeout
        self.ua = ua
        self._zamek = threading.Lock()
        self._posledni = {}
        self._cache_soubor = cache_soubor
        self._cache = {}
        self._zmenena = False
        self._obnova = threading.local()   # viz zapni_obnovu/vypni_obnovu
        if cache_soubor and os.path.exists(cache_soubor):
            try:
                with self._otevri(cache_soubor, "rt") as f:
                    self._cache = json.load(f)
            except Exception:
                self._cache = {}

    def _otevri(self, cesta, rezim):
        """Kes s priponou .gz se komprimuje (odpovedi rejstriku jsou objemne)."""
        if self._cache_soubor and self._cache_soubor.endswith(".gz"):
            return gzip.open(cesta, rezim, encoding="utf-8")
        return open(cesta, rezim, encoding="utf-8")

    def _cekej(self, url):
        host = urllib.parse.urlparse(url).netloc
        with self._zamek:
            posledni = self._posledni.get(host, 0.0)
            ted = time.monotonic()
            spat = self.prodleva - (ted - posledni)
            self._posledni[host] = max(ted, posledni + self.prodleva)
        if spat > 0:
            time.sleep(spat)

    def zapni_obnovu(self):
        """
        Pro aktualni vlakno vynuti pri ziskej() ignorovani cteni z kese (zapis
        do kese probiha porad, vysledek se tedy osvezi) - viz --obnovit-nenalezene,
        kde se takhle znovu dotazuji jen firmy se spatnym stavem z minula, aniz
        by se musela mazat cela kes.
        """
        self._obnova.aktivni = True

    def vypni_obnovu(self):
        self._obnova.aktivni = False

    def ziskej(self, url, hlavicky=None, json_body=None, ocisti=None, kontext=None):
        """
        Vrati telo odpovedi. `ocisti` je funkce nad rozparsovanym JSON, ktera
        z odpovedi vyhodi nepouzivane casti - kes pak neroste do stovek MB.
        `kontext` je volitelny ssl.SSLContext pro servery s neobvyklym
        certifikatem (viz TW_SSL_KONTEXT).
        """
        klic = url
        if json_body is not None:
            klic += "|" + json.dumps(json_body, sort_keys=True, ensure_ascii=False)
        if not getattr(self._obnova, "aktivni", False):
            with self._zamek:
                if klic in self._cache:
                    return self._cache[klic]

        h = {"User-Agent": self.ua, "Accept": "application/json"}
        if hlavicky:
            h.update(hlavicky)
        telo = None
        if json_body is not None:
            telo = json.dumps(json_body, ensure_ascii=False).encode("utf-8")
            h["Content-Type"] = "application/json"

        posledni_chyba = None
        for pokus in range(1, self.pokusy + 1):
            self._cekej(url)
            try:
                req = urllib.request.Request(url, data=telo, headers=h)
                with urllib.request.urlopen(req, timeout=self.timeout, context=kontext) as odp:
                    text = odp.read().decode("utf-8", errors="replace")
                if ocisti is not None:
                    try:
                        text = json.dumps(ocisti(json.loads(text)), ensure_ascii=False)
                    except (ValueError, KeyError, TypeError):
                        pass
                with self._zamek:
                    self._cache[klic] = text
                    self._zmenena = True
                return text
            except urllib.error.HTTPError as e:
                posledni_chyba = RuntimeError("HTTP %s" % e.code)
                if e.code == 404:
                    break
                if e.code in (429, 500, 502, 503, 504) and pokus < self.pokusy:
                    time.sleep(min(2 ** pokus, 8))
                    continue
                break
            except Exception as e:
                posledni_chyba = e
                if pokus < self.pokusy:
                    time.sleep(min(2 ** pokus, 8))
                    continue
        raise posledni_chyba or RuntimeError("neznama chyba")

    def zapomen(self, url, json_body=None):
        """
        Odstrani odpoved z kese. Pouziva se, kdyz server vratil HTTP 200
        s obsahem, ktery neni skutecna odpoved (napr. VIES pri privalu
        soubeznych dotazu vraci 200 s "MS_MAX_CONCURRENT_REQ" misto
        vysledku overeni) - bez tohoto by se docasna chyba ulozila do kese
        natrvalo, jako by to byla platna odpoved.
        """
        klic = url
        if json_body is not None:
            klic += "|" + json.dumps(json_body, sort_keys=True, ensure_ascii=False)
        with self._zamek:
            self._cache.pop(klic, None)

    def uloz_cache(self):
        if self._cache_soubor and self._zmenena:
            tmp = self._cache_soubor + ".tmp"
            with self._otevri(tmp, "wt") as f:
                json.dump(self._cache, f, ensure_ascii=False)
            os.replace(tmp, self._cache_soubor)


# ---------------------------------------------------------------------------
# Porovnavani nazvu firem
# ---------------------------------------------------------------------------

# NFKD tahle pismena nerozklada na zaklad + diakritiku - jsou to samostatna
# pismena, ne prekombinovane znaky ("ł" v "Spółka" tak bez tohoto prekladu
# zustane a rozbije tokenizaci nazvu).
BEZ_DIAKRITIKY_TABULKA = str.maketrans({
    "ł": "l", "Ł": "L", "đ": "d", "Đ": "D", "ø": "o", "Ø": "O",
    "þ": "th", "Þ": "Th", "æ": "ae", "Æ": "AE", "ß": "ss",
    "ı": "i", "İ": "I", "ĸ": "k", "ħ": "h", "Ħ": "H",
})


def bez_diakritiky(s):
    s = str(s).translate(BEZ_DIAKRITIKY_TABULKA)
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def normalizuj_nazev(s):
    """Male pismeno, bez diakritiky, bez pravni formy a interpunkce."""
    if not s:
        return ""
    s = bez_diakritiky(s).lower().replace("&", " and ")
    # zkratky typu "S.p.A." nebo "a.s." se jinak rozpadnou na osamocena
    # jednopismenna slova ("s", "p", "a") a neshodnou se s "spa"/"as" v PRAVNI_FORMY
    s = re.sub(r"([a-z0-9])\.(?=[a-z0-9])", r"\1", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    tokeny = [t for t in s.split() if t]
    for delka in (3, 2):
        while len(tokeny) > delka and " ".join(tokeny[-delka:]) in PRAVNI_FORMY:
            tokeny = tokeny[:-delka]
    while len(tokeny) > 1 and tokeny[-1] in PRAVNI_FORMY:
        tokeny = tokeny[:-1]
    zbytek = [t for t in tokeny if t not in PRAVNI_FORMY]
    return " ".join(zbytek or tokeny)


# Koncovky pravnich forem. Narozdil od PRAVNI_FORMY sem nepatri slova jako
# "group", "holding" nebo "company" - ta nesou cast identity firmy a bez nich
# by dotaz na rejstrik vratil desitky dcerinych spolecnosti.
PRAVNI_SUFIXY = {
    "as", "a s", "sro", "s r o", "spol s r o", "ks", "vos", "se", "ops", "zs",
    "akciova spolecnost", "akciova spolocnost", "spolecnost s rucenim omezenym",
    "spolocnost s rucenim obmedzenym",
    "gmbh", "mbh", "ag", "kg", "kgaa", "ohg", "ug", "eg", "gbr", "co kg", "se co kg",
    "aktiengesellschaft", "gesellschaft mit beschrankter haftung",
    "bv", "nv", "cv", "vof", "sa", "sas", "sasu", "sarl", "eurl",
    "srl", "spa", "snc", "sl", "slu", "sau", "s p a", "s a", "n v", "b v",
    "ltd", "limited", "plc", "llp", "llc", "inc", "incorporated", "corp", "corporation",
    "oy", "oyj", "ab", "aps", "asa", "kft", "zrt", "nyrt", "bt", "kkt",
    "sp z oo", "spzoo", "z oo", "doo", "dd", "ood", "eood", "ad", "ead",
    "lda", "kk", "yk", "pte", "pty", "bhd", "sdn", "dac", "cvba", "bvba",
    "sa nv", "nv sa", "a s ", "gmbh co kg", "se co kgaa",
}


def jadro_pro_hledani(s):
    """
    Nazev bez koncove pravni formy, jinak beze zmeny - "ORLEN S.A." -> "ORLEN",
    "Vodafone Group Plc" -> "Vodafone Group".

    Na dotazovani rejstriku se `normalizuj_nazev` pouzit neda: ta slouzi
    k porovnavani, vse zmensi a z "ORLEN S.A." udela "orlen s a", coz uz
    GLEIF nenajde.
    """
    tokeny = str(s or "").split()
    for _ in range(2):
        if len(tokeny) < 2:
            break
        posledni = re.sub(r"[^a-z]+", " ", bez_diakritiky(tokeny[-1]).lower()).strip()
        if posledni and posledni in PRAVNI_SUFIXY:
            tokeny = tokeny[:-1]
        else:
            break
    return re.sub(r"[\s,]+$", "", " ".join(tokeny)).strip() or str(s or "").strip()


def skore_syrove(a, b):
    """
    Podobnost nazvu bez odstraneni pravnich forem, jen bez diakritiky a
    interpunkce. Rozhoduje tam, kde `skore_shody` da remizu: "Vodafone Group
    Plc" sedi na "Vodafone Limited" i "Vodafone Group Public Limited Company"
    stejne, protoze "group" i "plc" se zahazuji - syrove skore uz rozdil vidi.
    """
    def uprav(s):
        return re.sub(r"[^a-z0-9]+", " ", bez_diakritiky(str(s or "")).lower()).strip()
    ua, ub = uprav(a), uprav(b)
    if not ua or not ub:
        return 0.0
    return round(SequenceMatcher(None, ua, ub).ratio(), 4)


def skore_shody(a, b):
    """0.0-1.0. Kombinace sekvencni podobnosti a prekryvu slov."""
    na, nb = normalizuj_nazev(a), normalizuj_nazev(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    sekvence = SequenceMatcher(None, na, nb).ratio()
    ta, tb = set(na.split()), set(nb.split())
    prunik = ta & tb
    dice = (2 * len(prunik)) / (len(ta) + len(tb))
    mensi, vetsi = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
    # Jeden nazev je slovnim podmnozinou druheho ("Kooperativa pojistovna" vs
    # "Kooperativa pojistovna, Vienna Insurance Group"). Bonus se udeluje jen
    # tehdy, kdyz je kratsi nazev dost specificky (aspon 2 slova), jinak by
    # "Apple" sedelo na "Apple Autos Properties".
    if mensi and mensi <= vetsi and len(mensi) >= 2:
        dice = max(dice, max(0.80, 0.97 - 0.015 * (len(vetsi) - len(mensi))))
    return round(max(sekvence, dice), 4)


# ---------------------------------------------------------------------------
# Vysledny zaznam
# ---------------------------------------------------------------------------

@dataclass
class Zaznam:
    hledany_nazev: str = ""
    jmeno: str = ""
    ulice: str = ""
    psc: str = ""
    mesto: str = ""
    zeme: str = ""
    ico: str = ""
    dic: str = ""
    nace: str = ""
    nace_popis: str = ""
    nace_vse: str = ""
    nace_zdroj: str = ""
    klasifikace: str = ""
    kod_kategorie: str = ""
    kategorie: str = ""
    skupina: str = ""
    zdroj_kategorie: str = ""
    zdroj: str = ""
    shoda: str = ""
    stav: str = STAV_NENALEZENO
    region: str = ""
    lei: str = ""
    reg_cislo: str = ""
    reg_rejstrik: str = ""
    rpo_id: str = ""    # interni ID v RPO SR pro dohledani SK NACE (viz doplnit_sk_nace)
    identifikator: str = ""
    pravni_forma: str = ""
    datum_vzniku: str = ""
    dic_overeno: str = ""
    odkaz: str = ""
    poznamka: str = ""
    obory: list = field(default_factory=list)      # QID oboru z Wikidat
    jmena: list = field(default_factory=list)      # dalsi nazvy (jine jazyky/prepisy)
    aktivni: bool = True
    kandidati: list = field(default_factory=list)
    nace_nejisty: bool = False   # jediny zapsany NACE je obecny/podpurny kod - viz PODPURNE_NACE
    nace_llm: str = ""           # NACE dohledany rucne/LLM (--export-nezarazene), pro srovnani


# ---------------------------------------------------------------------------
# ARES (Ceska republika)
# ---------------------------------------------------------------------------

# Cinnosti, ktere ma zapsanou temer kazda firma a hlavni obor neurcuji.
PODPURNE_NACE = {"00", "0000", "68", "682", "6820", "68200", "70", "702", "7010", "7022",
                 "74", "749", "7490", "77", "82", "829", "8299", "46", "461", "4619",
                 "469", "4690", "47", "471", "4778", "479", "4799", "52", "5210"}


def vyber_hlavni_nace(kody):
    """
    ARES vraci vsechny registrovane cinnosti bez oznaceni hlavni. Bereme prvni
    v poradi, ktera neni jen podpurna (pronajem, spravni cinnosti, obecny obchod).
    """
    if not kody:
        return ""
    hlavni = [k for k in kody if k not in PODPURNE_NACE]
    return (hlavni or kody)[0]


def _ares_ulice(sidlo):
    ulice = sidlo.get("nazevUlice") or sidlo.get("nazevCastiObce") or sidlo.get("nazevObce") or ""
    cp, co = sidlo.get("cisloDomovni"), sidlo.get("cisloOrientacni")
    cop = sidlo.get("cisloOrientacniPismeno") or ""
    if cp and co:
        cislo = "%s/%s%s" % (cp, co, cop)
    elif cp:
        cislo = str(cp)
    elif co:
        cislo = "%s%s" % (co, cop)
    else:
        cislo = ""
    return (ulice + " " + cislo).strip() or (sidlo.get("textovaAdresa") or "")


def _psc(hodnota):
    if hodnota is None:
        return ""
    p = re.sub(r"\s+", "", str(hodnota))
    return "%s %s" % (p[:3], p[3:]) if len(p) == 5 and p.isdigit() else p


def _ares_na_zaznam(s):
    sidlo = s.get("sidlo") or {}
    registrace = s.get("seznamRegistraci") or {}
    kody = s.get("czNace2008") or s.get("czNace") or []
    nace = vyber_hlavni_nace(kody)
    ico = s.get("ico") or ""
    return Zaznam(
        jmeno=s.get("obchodniJmeno") or "",
        ulice=_ares_ulice(sidlo),
        psc=_psc(sidlo.get("psc")),
        mesto=sidlo.get("nazevObce") or "",
        zeme=sidlo.get("kodStatu") or "CZ",
        ico=ico,
        dic=s.get("dic") or "",
        nace=nace,
        nace_popis=taxonomie.nazev_nace(nace),
        nace_vse=",".join(kody),
        region=sidlo.get("nazevKraje") or "",
        pravni_forma=s.get("pravniForma") or "",
        datum_vzniku=s.get("datumVzniku") or "",
        aktivni=not s.get("datumZaniku"),
        zdroj="ARES",
        odkaz="https://ares.gov.cz/ekonomicke-subjekty?ico=%s" % ico if ico else "",
        poznamka="; ".join(p for p in (
            "zanikl %s" % s.get("datumZaniku") if s.get("datumZaniku") else "",
            "registrovan k DPH, ARES neuvadi DIC"
            if registrace.get("stavZdrojeDph") == "AKTIVNI" and not s.get("dic") else "",
        ) if p),
    )


def ares_podle_ica(klient, ico):
    ico = re.sub(r"\D", "", str(ico)).zfill(8)
    return _ares_na_zaznam(json.loads(klient.ziskej(
        ARES_DETAIL.format(ico=ico), ocisti=_ares_ocisti)))


ARES_POLE = ("ico", "obchodniJmeno", "sidlo", "dic", "czNace2008", "czNace",
             "pravniForma", "datumVzniku", "datumZaniku", "seznamRegistraci")


def _ares_ocisti(data):
    """Z odpovedi ARES nechá jen pouzivana pole (dalsiUdaje byva radove vetsi)."""
    def zmensi(s):
        return {k: v for k, v in s.items() if k in ARES_POLE}
    if "ekonomickeSubjekty" in data:
        return {"ekonomickeSubjekty": [zmensi(s) for s in data["ekonomickeSubjekty"]]}
    if "zaznamy" in data:
        return {"zaznamy": [{k: v for k, v in z.items()
                             if k.startswith("czNacePrevazujici")} for z in data["zaznamy"]]}
    return zmensi(data)


def ares_prevazujici_nace(klient, ico):
    """
    Seznam czNace2008 je v ARES serazeny vzestupne, ne podle vyznamu. Zdroj RES
    ale vede prevazujici (hlavni) cinnost subjektu - tu pouzijeme prednostne.

    "00"/"0000" znamena u RES "neurceno" (subjekt nema prevazujici cinnost
    formalne nastavenou - typicky OSVC s vice zivnostmi bez oznacene hlavni)
    a neni to skutecny NACE kod - takovy zaznam se ignoruje, jinak by prepsal
    i dobre urcenou cinnost z vyber_hlavni_nace nepouzitelnou hodnotou.
    """
    ico = re.sub(r"\D", "", str(ico)).zfill(8)
    data = json.loads(klient.ziskej(ARES_RES.format(ico=ico), ocisti=_ares_ocisti))
    for zaznam in data.get("zaznamy", []):
        nace = zaznam.get("czNacePrevazujici2008") or zaznam.get("czNacePrevazujici")
        cislice = re.sub(r"\D", "", str(nace or ""))
        if cislice and cislice.strip("0"):
            return str(nace)
    return ""


def doplr_prevazujici_nace(klient, z):
    """Prepise hlavni NACE zaznamu prevazujici cinnosti z RES, je-li dostupna."""
    if z.zdroj != "ARES" or not z.ico:
        return
    try:
        nace = ares_prevazujici_nace(klient, z.ico)
    except Exception:
        return
    if nace:
        z.nace = nace
        z.nace_popis = taxonomie.nazev_nace(nace)


def _ares_vr_ocisti(data):
    """Z odpovedi Verejneho rejstriku (VR) nechá jen pole k dohledani vymazu."""
    def zmensi(z):
        return {k: v for k, v in z.items()
                if k in ("stavSubjektu", "datumVymazu", "pravniDuvodVymazu")}
    return {"zaznamy": [zmensi(z) for z in data.get("zaznamy", [])]}


def ares_vr_vymaz(klient, ico):
    """
    Doplnkovy dotaz do Verejneho rejstriku (zdroj VR). Narozdil od hlavniho
    ARES indexu (ekonomicke-subjekty), ktery jiz vymazane subjekty neobsahuje,
    VR drzi historii i po vymazu - vc. data a pravniho duvodu vymazu (napr.
    "z duvodu likvidace", "fuze se spolecnosti..."). Vraci None, pokud subjekt
    ve VR neni vubec veden, nebo pokud (zatim) vymazan neni.
    """
    ico = re.sub(r"\D", "", str(ico)).zfill(8)
    data = json.loads(klient.ziskej(ARES_VR.format(ico=ico), ocisti=_ares_vr_ocisti))
    for zaznam in data.get("zaznamy", []):
        datum = zaznam.get("datumVymazu")
        if not datum:
            continue
        duvody = [d.get("hodnota") for d in (zaznam.get("pravniDuvodVymazu") or [])
                  if d.get("hodnota")]
        return {"datum": datum, "duvod": "; ".join(duvody)}
    return None


def ares_podle_nazvu(klient, nazev, pocet=30):
    data = json.loads(klient.ziskej(ARES_HLEDAT, ocisti=_ares_ocisti, json_body={
        "obchodniJmeno": nazev, "pocet": min(pocet, 200), "start": 0}))
    return [_ares_na_zaznam(s) for s in data.get("ekonomickeSubjekty", [])]


# ---------------------------------------------------------------------------
# RPO - Registr pravnickych osob SR
# ---------------------------------------------------------------------------

def _sk_platny(polozky):
    """Z historie hodnot vrati aktualne platnou (bez validTo), jinak posledni."""
    if not polozky:
        return None
    aktualni = [p for p in polozky if not p.get("validTo")]
    return (aktualni or polozky)[-1]


def rpo_sk_podle_nazvu(klient, nazev, pocet=10):
    """
    Vyhledavaci pole RPO SR (fullName) je na interpunkci prekvapive citlive -
    s carkou pred pravni formou ("Firma, a.s.") nebo s teckovanou zkratkou
    ("Firma a.s." i "Firma a. s.") vraci 0 vysledku, zatimco bez pravni formy
    ("Firma") normalne najde. Dotaz se proto posila v ocistene podobe
    (normalizuj_nazev) - presnost vyberu spravneho zaznamu resi az nasledne
    porovnani skore_shody s puvodnim (neocistenym) nazvem.
    """
    dotaz = normalizuj_nazev(nazev) or nazev
    url = RPO_SK + "?" + urllib.parse.urlencode({"fullName": dotaz, "limit": pocet})
    data = json.loads(klient.ziskej(url, ocisti=lambda d: {"results": [
        {k: v for k, v in r.items()
         if k in ("id", "fullNames", "addresses", "identifiers", "establishment", "termination")}
        for r in d.get("results", [])]}))
    vysledky = []
    for r in data.get("results", []):
        jmeno = _sk_platny(r.get("fullNames")) or {}
        adresa = _sk_platny(r.get("addresses")) or {}
        ident = _sk_platny(r.get("identifiers")) or {}
        ico = str(ident.get("value") or "")
        psc = (adresa.get("postalCodes") or [""])[0]
        ulice = " ".join(x for x in (adresa.get("street"), adresa.get("buildingNumber")) if x)
        vysledky.append(Zaznam(
            jmeno=jmeno.get("value") or "",
            ulice=ulice,
            psc=_psc(psc),
            mesto=(adresa.get("municipality") or {}).get("value") or "",
            zeme="SK",
            ico=ico,
            dic="SK%s" % ico if ico else "",
            reg_cislo=ico,
            reg_rejstrik="RPO SR",
            rpo_id=str(r.get("id") or ""),
            datum_vzniku=r.get("establishment") or "",
            aktivni=not r.get("termination"),
            zdroj="RPO SR",
            odkaz="https://www.registeruz.sk/cruz-public/domain/accountingentity/simplesearch?ico=%s" % ico,
            poznamka="zanikl %s" % r.get("termination") if r.get("termination") else "",
        ))
    return vysledky


RPO_SK_DETAIL = "https://api.statistics.sk/rpo/v1/entity/{id}"


def _rpo_sk_detail_ocisti(data):
    return {k: v for k, v in data.items() if k in ("statisticalCodes", "legalForms")}


def rpo_sk_detail(klient, entity_id):
    """
    Vyhledavaci endpoint RPO SR (rpo_sk_podle_nazvu) vraci jen jmeno/adresu/ICO -
    hlavni ekonomicka cinnost (SK NACE, stejne cislovani jako CZ-NACE - obojí
    je narodni provedeni NACE Rev. 2) a pravni forma jsou az na urovni
    jednotliveho zaznamu (RPO SR ji vede jako soucast statistickych kodu,
    aktualizovanych Statistickym uradem SR).
    """
    data = json.loads(klient.ziskej(
        RPO_SK_DETAIL.format(id=entity_id), ocisti=_rpo_sk_detail_ocisti))
    hlavni = (data.get("statisticalCodes") or {}).get("mainActivity") or {}
    forma = _sk_platny(data.get("legalForms")) or {}
    return {
        "nace": str(hlavni.get("code") or ""),
        "pravni_forma": (forma.get("value") or {}).get("value") or "",
    }


def doplnit_sk_nace(klient, z):
    """Doplni SK NACE hlavni cinnosti a pravni formu z detailu RPO SR."""
    if z.zdroj != "RPO SR" or not z.rpo_id or z.nace:
        return
    try:
        detail = rpo_sk_detail(klient, z.rpo_id)
    except Exception:
        return
    if detail["nace"]:
        z.nace = detail["nace"]
        z.nace_popis = taxonomie.nazev_nace(detail["nace"])
    if detail["pravni_forma"] and not z.pravni_forma:
        z.pravni_forma = detail["pravni_forma"]


# ---------------------------------------------------------------------------
# GLEIF (LEI) - cely svet
# ---------------------------------------------------------------------------

GLEIF_RA = "https://api.gleif.org/api/v1/registration-authorities/{kod}"
GLEIF_ELF = "https://api.gleif.org/api/v1/entity-legal-forms/{kod}"
GLEIF_HLAVICKY = {"Accept": "application/vnd.api+json"}


def je_latinka(s):
    """True, pokud retezec nese aspon jedno pismeno a vsechna jsou latinkou."""
    pismena = [c for c in str(s or "") if c.isalpha()]
    if not pismena:
        return False
    return all("LATIN" in unicodedata.name(c, "") for c in pismena)


def _gleif_jmena(ent):
    """
    Vsechny nazvy entity. GLEIF vede pravni nazev v narodnim jazyce, takze
    korejska nebo japonska firma ma v `legalName` znaky, ktere se se vstupem
    nikdy neshodnou - anglicka varianta je az v `otherNames`. Bez ni by
    dodavatele z KR, JP, CN, TW nebo BG nesli dohledat vubec.
    """
    jmena = []
    hlavni = (ent.get("legalName") or {}).get("name") or ""
    if hlavni:
        jmena.append(hlavni)
    for klic in ("otherNames", "transliteratedOtherNames"):
        for o in ent.get(klic) or []:
            n = (o or {}).get("name")
            if n and n not in jmena:
                jmena.append(n)
    return jmena


def _gleif_adresa(ent):
    """
    Vrati adresu prednostne latinkou. Poradi: pravni sidlo -> alternativni
    jazykova varianta sidla -> centrala. Adresy typu "c/o ..." patri
    registracnimu agentovi, ne firme, takze se preskakuji.
    """
    def radky(adr):
        return [r for r in (adr.get("addressLines") or []) if r]

    moznosti = []
    for adr in (ent.get("legalAddress"), ent.get("headquartersAddress")):
        if adr:
            moznosti.append(adr)
    for adr in ent.get("otherAddresses") or []:
        if (adr.get("type") or "").startswith("ALTERNATIVE_LANGUAGE"):
            moznosti.append(adr)

    pouzitelne = [a for a in moznosti
                  if radky(a) and not re.match(r"(?i)\s*c/?o[\s.]", radky(a)[0])]
    if not pouzitelne:
        return ent.get("legalAddress") or {}, ""
    latinkou = [a for a in pouzitelne if je_latinka(radky(a)[0])]
    return (latinkou or pouzitelne)[0], ""


def _gleif_na_zaznam(rec):
    a = rec.get("attributes", {})
    ent = a.get("entity", {}) or {}
    jmena = _gleif_jmena(ent)
    # do vystupu jde nazev latinkou, aby byl seznam citelny; originalni zapis
    # zustava v poznamce
    zobrazit = next((j for j in jmena if je_latinka(j)), jmena[0] if jmena else "")
    puvodni = jmena[0] if jmena else ""

    adr, _ = _gleif_adresa(ent)
    radky = [r for r in (adr.get("addressLines") or []) if r]
    zeme = adr.get("country") or ent.get("jurisdiction") or ""
    stav = ent.get("status")
    reg = ent.get("registeredAs") or ""
    poznamky = []
    if stav and stav != "ACTIVE":
        poznamky.append("stav v LEI: %s" % stav)
    if puvodni and puvodni != zobrazit:
        poznamky.append("puvodni zapis nazvu: %s" % puvodni)
    return Zaznam(
        jmeno=zobrazit,
        ulice=radky[0] if radky else "",
        psc=adr.get("postalCode") or "",
        mesto=adr.get("city") or "",
        zeme=zeme[:2],
        reg_cislo=reg,
        reg_rejstrik=(ent.get("registeredAt") or {}).get("id") or "",
        region=adr.get("region") or "",
        lei=a.get("lei") or "",
        pravni_forma=(ent.get("legalForm") or {}).get("id") or "",
        datum_vzniku=(ent.get("creationDate") or "")[:10],
        jmena=jmena,
        aktivni=stav in (None, "", "ACTIVE"),
        zdroj="GLEIF",
        odkaz="https://search.gleif.org/#/record/%s" % a.get("lei", ""),
        poznamka="; ".join(poznamky),
    )


GLEIF_POLE = ("lei", "entity")


def _gleif_ocisti(data):
    """Z odpovedi GLEIF nechá jen atributy, ktere ctame."""
    def zmensi(r):
        return {"attributes": {k: v for k, v in (r.get("attributes") or {}).items()
                               if k in GLEIF_POLE}}
    if isinstance(data.get("data"), list):
        return {"data": [zmensi(r) for r in data["data"]]}
    if isinstance(data.get("data"), dict):
        return {"data": zmensi(data["data"])}
    return data


def gleif_podle_nazvu(klient, nazev, zeme=None, pocet=15):
    """
    Hleda v GLEIF nekolika zpusoby a vysledky slucuje.

    Jeden dotaz nestaci: filtr `entity.names` neprojde, kdyz nazev obsahuje
    pravni formu ("Vodafone Group Plc" nenajde nic, "Vodafone Group" ano),
    a naopak zkraceny nazev sam o sobe vraci desitky dcerinych spolecnosti.
    Posilame proto obe varianty a vyber nejlepsiho nechavame na skorovani.
    """
    nalezene, videne = [], set()

    def pridej(zaznamy):
        for z in zaznamy:
            if z.lei and z.lei not in videne:
                videne.add(z.lei)
                nalezene.append(z)

    def dotaz(parametry):
        url = GLEIF_API + "?" + urllib.parse.urlencode(parametry)
        data = json.loads(klient.ziskej(url, hlavicky=GLEIF_HLAVICKY, ocisti=_gleif_ocisti))
        return [_gleif_na_zaznam(r) for r in data.get("data", [])]

    jadro = jadro_pro_hledani(nazev)
    varianty = [nazev] + ([jadro] if jadro.lower() != nazev.lower() else [])

    for hledat in varianty:
        for s_zemi in ([True, False] if zeme else [False]):
            parametry = {"filter[entity.names]": hledat, "page[size]": min(pocet, 50)}
            if s_zemi:
                parametry["filter[entity.legalAddress.country]"] = zeme
            try:
                pridej(dotaz(parametry))
            except Exception:
                pass
        # dalsi varianty ma smysl vynechat, az kdyz uz mame presnou shodu
        # nazvu ze spravne zeme
        if any(z.zeme == zeme and max(skore_shody(nazev, j) for j in (z.jmena or [z.jmeno]))
               >= 0.97 for z in nalezene if z.jmena or z.jmeno):
            break

    # naseptavac vraci LEI presnych/blizkych nazvu, ktere fulltext casto minie
    if len(nalezene) < 3:
        try:
            url = GLEIF_AUTO + "?" + urllib.parse.urlencode({"field": "fulltext", "q": nazev})
            data = json.loads(klient.ziskej(url, hlavicky=GLEIF_HLAVICKY))
            leie = [d["relationships"]["lei-records"]["data"]["id"]
                    for d in data.get("data", [])[:5]
                    if d.get("relationships", {}).get("lei-records", {}).get("data")]
            for lei in leie:
                if lei in videne:
                    continue
                rec = json.loads(klient.ziskej(GLEIF_API + "/" + lei,
                                               hlavicky=GLEIF_HLAVICKY, ocisti=_gleif_ocisti))
                pridej([_gleif_na_zaznam(rec["data"])])
        except Exception:
            pass

    if not nalezene:
        zakladni = {"filter[fulltext]": nazev, "page[size]": min(pocet, 50)}
        if zeme:
            zakladni["filter[entity.legalAddress.country]"] = zeme
        try:
            pridej(dotaz(zakladni))
        except Exception:
            pass
    return nalezene


def gleif_popis_rejstriku(klient, kod):
    """Prelozi kod registracni autority (RA000657) na nazev rejstriku."""
    if not kod or not kod.startswith("RA"):
        return kod or ""
    try:
        data = json.loads(klient.ziskej(GLEIF_RA.format(kod=kod), hlavicky=GLEIF_HLAVICKY,
                                        ocisti=lambda d: {"data": {"attributes": {
                                            k: v for k, v in d["data"]["attributes"].items()
                                            if k in ("internationalName",
                                                     "internationalOrganizationName")}}}))
        a = data["data"]["attributes"]
        return ", ".join(x for x in (a.get("internationalOrganizationName"),
                                     a.get("internationalName")) if x) or kod
    except Exception:
        return kod


def gleif_popis_formy(klient, kod):
    """Prelozi ELF kod pravni formy (5RCH) na citelny nazev."""
    if not kod or len(kod) != 4:
        return kod or ""
    try:
        data = json.loads(klient.ziskej(GLEIF_ELF.format(kod=kod), hlavicky=GLEIF_HLAVICKY,
                                        ocisti=lambda d: {"data": {"attributes": {
                                            "names": d["data"]["attributes"].get("names")}}}))
        for n in data["data"]["attributes"].get("names") or []:
            nazev = n.get("transliteratedName") or n.get("localName")
            if nazev:
                return nazev
    except Exception:
        pass
    return kod


def gleif_podle_lei(klient, lei):
    """Primy dotaz na jeden LEI zaznam - presnejsi a rychlejsi nez hledani jmenem."""
    try:
        data = json.loads(klient.ziskej(GLEIF_API + "/" + lei, hlavicky=GLEIF_HLAVICKY,
                                        ocisti=_gleif_ocisti))
    except Exception:
        return None
    return _gleif_na_zaznam(data["data"]) if data.get("data") else None


def gleif_podle_registrovane(klient, cislo, zeme=None, pocet=10):
    """
    Presne vyhledani podle narodniho registracniho cisla (pole `registeredAs`
    v GLEIF - napr. nemecke HRB, slovenske ICO). Spolehlivejsi nez hledani
    jmenem, protoze cislo je (na rozdil od nazvu) jednoznacne.
    """
    parametry = {"filter[entity.registeredAs]": cislo, "page[size]": min(pocet, 25)}
    if zeme:
        parametry["filter[entity.legalAddress.country]"] = zeme
    url = GLEIF_API + "?" + urllib.parse.urlencode(parametry)
    try:
        data = json.loads(klient.ziskej(url, hlavicky=GLEIF_HLAVICKY, ocisti=_gleif_ocisti))
    except Exception:
        return []
    vysledky = [_gleif_na_zaznam(r) for r in data.get("data", [])]
    # filtr je u GLEIF niekdy jen fulltextovy - overit presnou shodu cisla
    cislo_n = re.sub(r"[\s.\-]", "", cislo).upper()
    presne = [z for z in vysledky if re.sub(r"[\s.\-]", "", z.reg_cislo).upper() == cislo_n]
    return presne or vysledky


# ---------------------------------------------------------------------------
# SEC EDGAR - USA
# ---------------------------------------------------------------------------

def _bez_ns(prvek, tag):
    """
    Najde primeho potomka podle mistniho jmena znacky bez ohledu na jmenny
    prostor. Atom feed EDGARu vse zabaluje do namespace
    "http://www.w3.org/2005/Atom", takze holé `Element.find("cik")` na nem
    nikdy nic nenajde - tise vraci None i tam, kde element skutecne je.
    """
    return next((e for e in prvek if e.tag.rsplit("}", 1)[-1] == tag), None)


def _vsichni_bez_ns(prvek, tag):
    return [e for e in prvek.iter() if e.tag.rsplit("}", 1)[-1] == tag]


def _edgar_adresa(prvek):
    for typ in ("business", "mailing"):
        for adr in _vsichni_bez_ns(prvek, "address"):
            if adr.get("type") != typ:
                continue

            def h(tag):
                e = _bez_ns(adr, tag)
                return (e.text or "").strip() if e is not None and e.text else ""
            ulice = " ".join(x for x in (h("street1"), h("street2")) if x)
            if ulice or h("city"):
                return ulice, h("city"), h("state"), h("zip")
    return "", "", "", ""


def _edgar_na_zaznam(ci, cik_zaloha="", nazev_zaloha=""):
    def h(tag):
        e = _bez_ns(ci, tag)
        return (e.text or "").strip() if e is not None and e.text else ""
    cik = h("cik") or cik_zaloha
    sic = h("assigned-sic")
    # v USA se NACE nepouziva - domaci obdoba je NAICS (SIC u SEC je jeho
    # starsi predchudce). NACE se dopocitava jen jako priblizny ekvivalent
    # pro zarazeni do vlastni taxonomie.
    nace = taxonomie.sic_na_nace(sic) or ""
    naics = taxonomie.sic_na_naics(sic) or ""
    ulice, mesto, stat, psc = _edgar_adresa(ci)
    cik_cislo = cik.lstrip("0") if cik else ""
    return Zaznam(
        jmeno=h("conformed-name") or nazev_zaloha,
        ulice=ulice, psc=psc, mesto=mesto, region=stat, zeme="US",
        reg_cislo=cik_cislo, reg_rejstrik="SEC CIK",
        nace=nace, nace_popis=taxonomie.nazev_nace(nace),
        klasifikace="NAICS %s - %s" % (naics, taxonomie.nazev_naics(naics)) if naics else "",
        zdroj="SEC EDGAR",
        odkaz="https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=%s" % cik if cik else "",
        poznamka=("SIC %s %s" % (sic, h("assigned-sic-desc"))).strip() if sic else "",
    )


def edgar_podle_nazvu(klient, nazev, pocet=10):
    url = EDGAR_API + "?" + urllib.parse.urlencode({
        "company": nazev, "type": "", "dateb": "", "owner": "include",
        "count": pocet, "action": "getcompany", "output": "atom"})
    try:
        koren = ET.fromstring(klient.ziskej(
            url, hlavicky={"Accept": "application/atom+xml"}).encode("utf-8"))
    except (ET.ParseError, Exception):
        return []

    ci = _bez_ns(koren, "company-info")
    if ci is not None:
        return [_edgar_na_zaznam(ci)]

    vysledky = []
    for entry in koren.iter():
        if not entry.tag.endswith("entry"):
            continue
        # u vice shodujicich se firem je <company-info> vnorene v <content>,
        # ne primo pod <entry> - proto hledani do hloubky
        ci = next(iter(_vsichni_bez_ns(entry, "company-info")), None)
        if ci is not None:
            vysledky.append(_edgar_na_zaznam(ci))
            continue
        titul = next((e for e in entry if e.tag.endswith("title")), None)
        if titul is not None and titul.text:
            m = re.match(r"(.*?)\s*\(CIK (\d+)\)", titul.text.strip())
            if m:
                vysledky.append(_edgar_na_zaznam(entry, m.group(2), m.group(1)))
    for z in vysledky:
        if not z.jmeno:
            # U hromadneho vyhledavani podle jmena SEC u firem s vic
            # registrovanymi subjekty (napr. vic dcerinych spolecnosti
            # jedne skupiny) vraci zaznam bez <conformed-name> vubec -
            # bez jmena by kandidat dostal 0% shodu a jinak spravny
            # CIK/SIC (napr. skutecny obor cinnosti) by se tise zahodil.
            z.jmeno = nazev
    return vysledky


def edgar_podle_cik(klient, cik):
    """Primy dotaz na jedno CIK - presnejsi a rychlejsi nez hledani jmenem."""
    cik_cislo = re.sub(r"\D", "", str(cik))
    if not cik_cislo:
        return None
    url = EDGAR_API + "?" + urllib.parse.urlencode({
        "CIK": cik_cislo, "type": "", "dateb": "", "owner": "include",
        "count": 1, "action": "getcompany", "output": "atom"})
    try:
        koren = ET.fromstring(klient.ziskej(
            url, hlavicky={"Accept": "application/atom+xml"}).encode("utf-8"))
    except (ET.ParseError, Exception):
        return None
    ci = _bez_ns(koren, "company-info")
    return _edgar_na_zaznam(ci) if ci is not None else None


# ---------------------------------------------------------------------------
# Nemecko - Handelsregister pres lokalni kopii OffeneRegister.de
#
# GLEIF obsahuje jen firmy s LEI (povinne hlavne pro ucastniky financnich
# trhu), takze bezna mala nemecka GmbH/UG v nem typicky vubec neni. Nemecko
# nema oficialni verejne API k Handelsregisteru - jedina otevrena alternativa
# je bulk export OpenCorporates zverejnovany projektem OffeneRegister.de
# (OKF Deutschland). Jeho zive dotazovaci API (db.offeneregister.de) je
# dlouhodobe nedostupne (padly backend), proto se pouziva primo stazitelna
# SQLite kopie s FTS5 indexem (daten.offeneregister.de) - viz
# de_pripravit_databazi(). Data jsou sbirana do zacatku 2019 a dal se
# neaktualizuji, takze u aktivity/noveho jednatele pocitejte s tim, ze
# nemusi byt aktualni - na rozdil od ARES/INSEE tu nejde o zivy rejstrik.
# ---------------------------------------------------------------------------

DE_REGISTER_URL = "https://daten.offeneregister.de/openregister.db.gz"
DE_REGISTER_DB = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               "de_handelsregister.db")

# Serazeno od nejdelsich k nejkratsim, aby se napr. "GmbH & Co. KG" nerozpoznalo
# jen jako "KG".
DE_PRAVNI_FORMY = (
    "UG (haftungsbeschränkt) & Co. KG", "GmbH & Co. KGaA", "GmbH & Co. KG",
    "AG & Co. KG", "UG (haftungsbeschränkt)", "gGmbH", "GmbH", "KGaA",
    "OHG", "PartG", "GbR", "e.K.", "eK", "e.G.", "eG", "e.V.", "eV",
    "mbH", "AG", "KG", "SE",
)

_DE_LOCAL = threading.local()


def de_pripravit_databazi(force=False):
    """
    Stahne (~740 MB) a rozbali (~2,6 GB) lokalni SQLite kopii nemeckeho
    Handelsregisteru z daten.offeneregister.de (OpenCorporates/OKF
    Deutschland, licence CC BY 4.0). Jednorazova priprava pred prvnim
    pouzitim --bez-de neaktivniho behu.
    """
    if os.path.exists(DE_REGISTER_DB) and not force:
        print("%s uz existuje, preskakuji stahovani (spustte se --force pro "
              "znovustazeni)" % DE_REGISTER_DB, file=sys.stderr)
        return
    gz_cesta = DE_REGISTER_DB + ".gz"
    print("Stahuji %s (~740 MB)..." % DE_REGISTER_URL, file=sys.stderr)
    with urllib.request.urlopen(DE_REGISTER_URL, timeout=120) as odpoved, \
            open(gz_cesta, "wb") as f:
        while True:
            blok = odpoved.read(1024 * 1024)
            if not blok:
                break
            f.write(blok)
    print("Rozbaluji do %s (~2,6 GB)..." % DE_REGISTER_DB, file=sys.stderr)
    with gzip.open(gz_cesta, "rb") as f_in, open(DE_REGISTER_DB, "wb") as f_out:
        while True:
            blok = f_in.read(1024 * 1024)
            if not blok:
                break
            f_out.write(blok)
    os.remove(gz_cesta)
    print("Hotovo -> %s" % DE_REGISTER_DB, file=sys.stderr)


def _de_pripojeni():
    if not os.path.exists(DE_REGISTER_DB):
        raise RuntimeError(
            "chybi %s - spustte 'python3 dodavatele.py --pripravit-de-rejstrik'"
            % os.path.basename(DE_REGISTER_DB))
    spojeni = getattr(_DE_LOCAL, "spojeni", None)
    if spojeni is None:
        spojeni = sqlite3.connect(
            "file:%s?mode=ro" % DE_REGISTER_DB, uri=True, check_same_thread=False)
        _DE_LOCAL.spojeni = spojeni
    return spojeni


def _de_pravni_forma(jmeno):
    for forma in DE_PRAVNI_FORMY:
        if jmeno.endswith(forma):
            return forma
    return ""


# "Ulice cislo, PSC Mesto." - format adres v datech OpenCorporates. Cast
# firem ma adresu neuplnou (jen ulice, bez PSC/mesta) - v tom pripade zustane
# cely retezec v ulici a psc/mesto prazdne.
_DE_ADRESA_RE = re.compile(r"^(?P<ulice>.*?),\s*(?P<psc>\d{5})\s+(?P<mesto>.*?)\.?\s*$")


def _de_adresa(retezec):
    m = _DE_ADRESA_RE.match((retezec or "").strip())
    if not m:
        return (retezec or "").rstrip("., ").strip(), "", ""
    return m.group("ulice").strip(), m.group("psc"), m.group("mesto").strip()


_DE_SLOUPCE = ("company_number", "name", "registered_address", "current_status",
               "register_art", "register_nummer")


def _de_na_zaznam(radek):
    company_number, name, registered_address, current_status, register_art, \
        register_nummer = radek
    ulice, psc, mesto = _de_adresa(registered_address)
    aktivni = current_status == "currently registered"
    return Zaznam(
        jmeno=name or "",
        ulice=ulice, psc=psc, mesto=mesto, zeme="DE",
        reg_cislo=("%s %s" % (register_art, register_nummer)).strip()
                  or company_number,
        reg_rejstrik="Handelsregister",
        pravni_forma=_de_pravni_forma(name or ""),
        aktivni=aktivni,
        zdroj="OffeneRegister.de (data k 2019)",
        poznamka="vymazana/zanikla firma (Handelsregister, stav k 2019)"
                 if not aktivni else "",
    )


def de_podle_nazvu(klient, nazev, pocet=15):
    """
    Fulltextove hledani v lokalni kopii Handelsregisteru (FTS5, ~5,3 mil.
    firem vsech velikosti vc. malych GmbH/UG bez LEI).
    """
    dotaz = " ".join(re.findall(r"\w+", nazev, re.UNICODE))
    if not dotaz:
        return []
    spojeni = _de_pripojeni()
    kurzor = spojeni.execute(
        "SELECT c.%s FROM company_fts f JOIN company c ON c.id = f.rowid "
        "WHERE company_fts MATCH ? ORDER BY bm25(company_fts) LIMIT ?"
        % ", c.".join(_DE_SLOUPCE),
        (dotaz, pocet))
    return [_de_na_zaznam(r) for r in kurzor.fetchall()]


def de_podle_registru(register_art, register_nummer):
    """
    Presny dotaz na cislo zapisu (napr. HRB 150148). Cislo samo o sobe neni
    jednoznacne - stejne cislo pouzivaji ruzne rejstrikove soudy - proto se
    (stejne jako u GLEIF narodniho cisla) muze vratit vic kandidatu k
    rozliseni podle nazvu/adresy.
    """
    spojeni = _de_pripojeni()
    kurzor = spojeni.execute(
        "SELECT c.%s FROM company c WHERE c.register_art = ? "
        "AND c.register_nummer = ?" % ", c.".join(_DE_SLOUPCE),
        (register_art, register_nummer))
    return [_de_na_zaznam(r) for r in kurzor.fetchall()]


_DE_REG_CISLO_RE = re.compile(r"\b(HRA|HRB|GnR|PR|VR)\s*0*(\d+)\b", re.IGNORECASE)


def de_rozloz_reg_cislo(text):
    """Vytahne (druh, cislo) z retezce jako 'HRB 150148' nebo 'HRB150148'."""
    m = _DE_REG_CISLO_RE.search(text or "")
    return (m.group(1).upper(), m.group(2)) if m else (None, None)


# ---------------------------------------------------------------------------
# Nemecko - OpenRegister.de API (volitelne, vyzaduje vlastni API klic)
#
# Na rozdil od lokalni kopie Handelsregisteru (vyse) tohle je placene/
# kreditove API tretí strany, ktere ale narozdil od samotneho Handelsregisteru
# vede i skutecny obor cinnosti (WZ2025 - nemecka obdoba NACE) a text
# "Gegenstand des Unternehmens". Vyzaduje vlastni ucet a API klic
# (openregister.de, zdarma 500 kreditu/mesic, 10 kreditu/dotaz na detail) -
# klic se NIKDY neuklada v kodu ani v repozitari, jen se preda pri behu
# (--de-api-klic, nebo promenna prostredi OPENREGISTER_API_KEY).
# ---------------------------------------------------------------------------

OPENREGISTER_API = "https://api.openregister.de"
OPENREGISTER_AUTOCOMPLETE = OPENREGISTER_API + "/v1/autocomplete/company"
OPENREGISTER_DETAIL = OPENREGISTER_API + "/v1/company/{id}"


def _openregister_adresa(adresa):
    adresa = adresa or {}
    ulice = " ".join(x for x in (adresa.get("street"),) if x)
    return ulice, adresa.get("postal_code") or "", adresa.get("city") or ""


def _openregister_na_zaznam(polozka):
    ulice, psc, mesto = _openregister_adresa(polozka.get("address"))
    return Zaznam(
        jmeno=polozka.get("name") or "",
        ulice=ulice, psc=_psc(psc), mesto=mesto, zeme=polozka.get("country") or "DE",
        reg_cislo=("%s %s" % (polozka.get("register_type") or "",
                              polozka.get("register_number") or "")).strip(),
        reg_rejstrik="Handelsregister (%s)" % polozka.get("register_court")
                    if polozka.get("register_court") else "Handelsregister",
        pravni_forma=polozka.get("legal_form") or "",
        aktivni=bool(polozka.get("active", True)),
        zdroj="OpenRegister.de",
        odkaz="",
        poznamka="",
        identifikator=polozka.get("company_id") or "",   # docasne - viz doplnit_openregister_nace
    )


def _openregister_detail_ocisti(data):
    return {k: v for k, v in data.items() if k in ("industry_codes", "purpose", "purposes")}


def openregister_podle_nazvu(klient, nazev, api_klic, pocet=15):
    """
    Vyhledani podle jmena pres OpenRegister.de (autocomplete) - neobsahuje
    jeste WZ kod (ten je az v detailu jednotlive firmy, viz
    doplnit_openregister_nace), ale uz obsahuje adresu, pravni formu a text
    predmetu podnikani (purpose).
    """
    if not api_klic:
        return []
    url = OPENREGISTER_AUTOCOMPLETE + "?" + urllib.parse.urlencode({"query": nazev})
    data = json.loads(klient.ziskej(
        url, hlavicky={"Authorization": "Bearer %s" % api_klic}))
    vysledky = []
    for r in (data.get("results") or [])[:pocet]:
        z = _openregister_na_zaznam(r)
        z.poznamka = (r.get("purpose") or "")[:300]
        vysledky.append(z)
    return vysledky


def doplnit_openregister_nace(klient, z, api_klic):
    """
    Dotahne WZ2025 kod (nemecka obdoba NACE) pro uz vybranou nejlepsi shodu -
    autocomplete vyhledavani ho nevraci, je az v detailu jedne konkretni
    firmy (company_id). Vola se jen jednou, po vyberu nejlepsiho kandidata -
    ne pro kazdy vraceny kandidat zvlast, aby se zbytecne neplytvalo kredity.
    """
    if z.zdroj != "OpenRegister.de" or not z.identifikator or z.nace:
        return
    url = OPENREGISTER_DETAIL.format(id=urllib.parse.quote(z.identifikator, safe=""))
    data = json.loads(klient.ziskej(
        url, hlavicky={"Authorization": "Bearer %s" % api_klic},
        ocisti=_openregister_detail_ocisti))
    kody = ((data.get("industry_codes") or {}).get("WZ2025") or [])
    if kody:
        kod = re.sub(r"\D", "", str(kody[0].get("code") or ""))
        if kod:
            z.nace = kod
            z.nace_popis = taxonomie.nazev_nace(kod)
            z.nace_vse = ",".join(sorted({
                c for c in (re.sub(r"\D", "", str(k.get("code") or "")) for k in kody) if c}))
            z.nace_zdroj = "WZ2025 (OpenRegister.de)"


# ---------------------------------------------------------------------------
# Pobalti + Svedsko/Finsko - Scoris API (volitelne, vyzaduje vlastni API klic)
#
# scoris.eu (ne zamenovat s ceskym/litevskym scoris.lt - jine API, jiny klic)
# pokryva jen SE/FI/EE/LV/LT/GB - u GB uz mame lepsi bezplatny zdroj
# (Companies House), proto se zde pouziva jen pro SE/FI/EE/LV/LT. Vyhledavani
# jmenem nevraci adresu ani NACE, jen jmeno+zemi+registracni cislo - detail
# (a tim i skutecny NACE) se dotahuje az pro jiz vybraneho nejlepsiho
# kandidata, aby se neplytvalo kredity na kandidaty, kteri nakonec nejsou
# vybrani.
# ---------------------------------------------------------------------------

SCORIS_API = "https://scoris.eu"
SCORIS_SEARCH = SCORIS_API + "/api/v1/company-search/"
SCORIS_DETAIL = SCORIS_API + "/api/v1/company/{zeme}/{regcode}/"
SCORIS_ZEME = {"SE", "FI", "EE", "LV", "LT"}


def scoris_podle_nazvu(klient, nazev, api_klic, zeme=None, pocet=15):
    if not api_klic:
        return []
    parametry = {"name": nazev, "limit": min(pocet, 100)}
    if zeme:
        parametry["country_code"] = zeme
    url = SCORIS_SEARCH + "?" + urllib.parse.urlencode(parametry)
    data = json.loads(klient.ziskej(url, hlavicky={"X-API-Key": api_klic}))
    return [
        Zaznam(
            jmeno=r.get("name") or "",
            zeme=r.get("country_code") or "",
            reg_cislo=r.get("regcode") or "",
            reg_rejstrik="obchodni rejstrik (Scoris)",
            zdroj="Scoris",
            identifikator="%s:%s" % (r.get("country_code"), r.get("regcode")),
        )
        for r in (data or [])[:pocet]
    ]


def _scoris_detail_ocisti(data):
    return {k: v for k, v in data.items() if k in ("company", "meta")}


def doplnit_scoris_detail(klient, z, api_klic):
    """
    Dotahne adresu, pravni formu, DIC a skutecny NACE pro uz vybraneho
    nejlepsiho kandidata - vyhledavani jmenem (scoris_podle_nazvu) samo
    o sobe vraci jen jmeno/zemi/registracni cislo.
    """
    if z.zdroj != "Scoris" or not z.identifikator or z.pravni_forma:
        return
    zeme, regcode = z.identifikator.split(":", 1)
    url = SCORIS_DETAIL.format(zeme=zeme, regcode=urllib.parse.quote(regcode, safe=""))
    data = json.loads(klient.ziskej(
        url, hlavicky={"X-API-Key": api_klic}, ocisti=_scoris_detail_ocisti))
    spol = data.get("company") or {}
    adresa = spol.get("address") or {}
    z.psc = adresa.get("postal_code") or z.psc
    casti = (adresa.get("address") or "").split(",")
    z.ulice = casti[0].strip()
    if len(casti) > 1:
        # posledni cast bývá "PSC MESTO" nebo jen "MESTO" - admin_name1/2
        # jsou kraj/region, ne mesto (napr. FI "Uusimaa" pro Espoo)
        posledni = casti[-1].strip()
        if z.psc and posledni.startswith(z.psc):
            posledni = posledni[len(z.psc):].strip()
        z.mesto = posledni or adresa.get("admin_name2") or adresa.get("admin_name1") or z.mesto
    else:
        z.mesto = adresa.get("admin_name2") or adresa.get("admin_name1") or z.mesto
    z.pravni_forma = spol.get("type") or "neurcena"
    z.dic = z.dic or spol.get("vat_code") or ""
    nace = ((spol.get("classifications") or {}).get("nace") or {})
    kod = re.sub(r"\D", "", str(nace.get("nace_code") or ""))
    if kod:
        z.nace = kod
        z.nace_popis = taxonomie.nazev_nace(kod)
        z.nace_vse = kod
        z.nace_zdroj = "NACE (Scoris)"


# ---------------------------------------------------------------------------
# Velka Britanie - Companies House (bezplatny bulk export, zadna registrace)
#
# Na rozdil od nemeckeho Handelsregisteru obsahuje bulk soubor primo i obor
# cinnosti (UK SIC 2007 - stejna urovnova struktura jako NACE Rev. 2, prvni
# 2 cislice = divize se stejnym vyznamem), takze pro UK neni potreba fallback
# na Wikidata jen kvuli oboru - jen na samotne dohledani firmy. Soubor se
# aktualizuje mesicne (nahran kolem zacatku mesice), zadny API klic ani
# registrace neni potreba - viz http://download.companieshouse.gov.uk/en_output.html
# ---------------------------------------------------------------------------

GB_REGISTER_URL_VZOR = "http://download.companieshouse.gov.uk/BasicCompanyDataAsOneFile-%s-01.zip"
GB_REGISTER_DB = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               "gb_companies_house.db")

# (nazev sloupce v CSV, nazev sloupce v nasi SQLite tabulce)
_GB_SLOUPCE_CSV = (
    ("CompanyNumber", "company_number"), ("CompanyName", "name"),
    ("CompanyStatus", "status"), ("CompanyCategory", "category"),
    ("IncorporationDate", "incorporation_date"),
    ("RegAddress.AddressLine1", "address1"), ("RegAddress.AddressLine2", "address2"),
    ("RegAddress.PostTown", "post_town"), ("RegAddress.PostCode", "post_code"),
    ("SICCode.SicText_1", "sic1"), ("SICCode.SicText_2", "sic2"),
    ("SICCode.SicText_3", "sic3"), ("SICCode.SicText_4", "sic4"),
)
_GB_SLOUPCE = tuple(nazev_db for _, nazev_db in _GB_SLOUPCE_CSV)

_GB_LOCAL = threading.local()


def _gb_najdi_aktualni_url():
    """Soubor je pojmenovany podle mesice publikace a bezici mesic jeste
    nemusi byt nahrany - zkusi aktualni mesic, pak jeden zpet."""
    ted = time.gmtime()
    rok, mesic = ted.tm_year, ted.tm_mon
    for _ in range(2):
        url = GB_REGISTER_URL_VZOR % ("%04d-%02d" % (rok, mesic))
        try:
            urllib.request.urlopen(
                urllib.request.Request(url, method="HEAD"), timeout=20).close()
            return url
        except urllib.error.HTTPError:
            mesic -= 1
            if mesic <= 0:
                mesic, rok = 12, rok - 1
    raise RuntimeError(
        "aktualni soubor Companies House se nepodarilo najit (zkuseny posledni "
        "2 mesice) - zkontrolujte http://download.companieshouse.gov.uk/en_output.html")


def _gb_importuj_csv(f, cesta_db):
    """Naimportuje bulk CSV do nove SQLite s FTS5 indexem nad nazvem firmy.
    FTS index se stavi az po hromadnem vlozeni dat (rebuild) - podstatne
    rychlejsi nez prubezna udrzba indexu pri ~5 mil. radcich."""
    if os.path.exists(cesta_db):
        os.remove(cesta_db)
    spojeni = sqlite3.connect(cesta_db)
    spojeni.execute("PRAGMA synchronous=OFF")
    spojeni.execute("PRAGMA journal_mode=MEMORY")
    spojeni.execute("CREATE TABLE company (id INTEGER PRIMARY KEY, %s)"
                    % ", ".join("%s TEXT" % s for s in _GB_SLOUPCE))

    cteni = csv.reader(f)
    hlavicka = [h.strip() for h in next(cteni)]
    try:
        indexy = [hlavicka.index(nazev_csv) for nazev_csv, _ in _GB_SLOUPCE_CSV]
    except ValueError as e:
        raise RuntimeError(
            "Companies House CSV nema ocekavany sloupec (%s) - format souboru "
            "se zjevne zmenil" % e)

    vlozit = "INSERT INTO company (%s) VALUES (%s)" % (
        ", ".join(_GB_SLOUPCE), ", ".join("?" * len(_GB_SLOUPCE)))
    davka = []
    with spojeni:
        for radek in cteni:
            davka.append(tuple(radek[i] if i < len(radek) else "" for i in indexy))
            if len(davka) >= 5000:
                spojeni.executemany(vlozit, davka)
                davka.clear()
        if davka:
            spojeni.executemany(vlozit, davka)

    spojeni.execute(
        "CREATE VIRTUAL TABLE company_fts USING FTS5(name, content='company', "
        "content_rowid='id')")
    spojeni.execute("INSERT INTO company_fts(company_fts) VALUES ('rebuild')")
    spojeni.commit()
    spojeni.close()


def gb_pripravit_databazi(force=False):
    """
    Stahne mesicni bulk export Companies House (~500 MB zip, ~5 mil. firem
    vc. SIC/oboru cinnosti) a naimportuje ho do lokalni SQLite s FTS5 indexem -
    zadny API klic ani registrace neni potreba.
    """
    if os.path.exists(GB_REGISTER_DB) and not force:
        print("%s uz existuje, preskakuji stahovani (smazte soubor a spustte "
              "znovu pro aktualizaci)" % GB_REGISTER_DB, file=sys.stderr)
        return
    url = _gb_najdi_aktualni_url()
    zip_cesta = GB_REGISTER_DB + ".zip"
    print("Stahuji %s (~500 MB)..." % url, file=sys.stderr)
    with urllib.request.urlopen(url, timeout=120) as odpoved, open(zip_cesta, "wb") as f:
        while True:
            blok = odpoved.read(1024 * 1024)
            if not blok:
                break
            f.write(blok)
    print("Rozbaluji a importuji do %s..." % GB_REGISTER_DB, file=sys.stderr)
    with zipfile.ZipFile(zip_cesta) as z:
        nazev_csv = next(n for n in z.namelist() if n.lower().endswith(".csv"))
        with z.open(nazev_csv) as f_bin:
            _gb_importuj_csv(io.TextIOWrapper(f_bin, encoding="utf-8", errors="replace"),
                             GB_REGISTER_DB)
    os.remove(zip_cesta)
    print("Hotovo -> %s" % GB_REGISTER_DB, file=sys.stderr)


def _gb_pripojeni():
    if not os.path.exists(GB_REGISTER_DB):
        raise RuntimeError(
            "chybi %s - spustte 'python3 dodavatele.py --pripravit-gb-rejstrik'"
            % os.path.basename(GB_REGISTER_DB))
    spojeni = getattr(_GB_LOCAL, "spojeni", None)
    if spojeni is None:
        spojeni = sqlite3.connect(
            "file:%s?mode=ro" % GB_REGISTER_DB, uri=True, check_same_thread=False)
        _GB_LOCAL.spojeni = spojeni
    return spojeni


_GB_SIC_RE = re.compile(r"^\s*(\d{4,5})")


def _gb_nace_ze_sic(sic_text):
    """'62020 - Information technology consultancy activities' -> '62020'."""
    m = _GB_SIC_RE.match(sic_text or "")
    return m.group(1) if m else ""


def _gb_na_zaznam(radek):
    (company_number, name, status, category, incorporation_date,
     address1, address2, post_town, post_code, sic1, sic2, sic3, sic4) = radek
    ulice = ", ".join(x for x in (address1, address2) if x)
    nace = _gb_nace_ze_sic(sic1)
    nace_vse = ",".join(sorted({c for c in (
        _gb_nace_ze_sic(sic1), _gb_nace_ze_sic(sic2),
        _gb_nace_ze_sic(sic3), _gb_nace_ze_sic(sic4)) if c}))
    aktivni = (status or "").strip().lower() == "active"
    return Zaznam(
        jmeno=name or "",
        ulice=ulice, psc=post_code or "", mesto=post_town or "", zeme="GB",
        reg_cislo=company_number or "",
        reg_rejstrik="Companies House",
        pravni_forma=category or "",
        nace=nace,
        nace_popis=taxonomie.nazev_nace(nace),
        nace_vse=nace_vse,
        nace_zdroj="UK SIC 2007 (Companies House)" if nace else "",
        datum_vzniku=incorporation_date or "",
        aktivni=aktivni,
        zdroj="Companies House",
        odkaz="https://find-and-update.company-information.service.gov.uk/company/%s"
              % company_number if company_number else "",
        poznamka="" if aktivni else "stav v Companies House: %s" % status,
    )


def gb_podle_nazvu(klient, nazev, pocet=15):
    """Fulltextove hledani v lokalni kopii Companies House (FTS5, ~5 mil. firem)."""
    dotaz = " ".join(re.findall(r"\w+", nazev, re.UNICODE))
    if not dotaz:
        return []
    spojeni = _gb_pripojeni()
    kurzor = spojeni.execute(
        "SELECT c.%s FROM company_fts f JOIN company c ON c.id = f.rowid "
        "WHERE company_fts MATCH ? ORDER BY bm25(company_fts) LIMIT ?"
        % ", c.".join(_GB_SLOUPCE),
        (dotaz, pocet))
    return [_gb_na_zaznam(r) for r in kurzor.fetchall()]


def gb_podle_cisla(company_number):
    """Presny dotaz na registracni cislo (Company Number) - jednoznacne, na
    rozdil od nemeckeho HRB/HRA cislo v UK neni sdilene mezi ruznymi soudy."""
    spojeni = _gb_pripojeni()
    kurzor = spojeni.execute(
        "SELECT c.%s FROM company c WHERE c.company_number = ?"
        % ", c.".join(_GB_SLOUPCE),
        (company_number.strip().upper(),))
    radek = kurzor.fetchone()
    return _gb_na_zaznam(radek) if radek else None


# ---------------------------------------------------------------------------
# Francie - Recherche d'entreprises (INSEE/INPI, data.gouv.fr)
# ---------------------------------------------------------------------------

def fr_dic_ze_siren(siren):
    """
    Francouzske DIC se pocita ze SIREN kontrolnim vzorcem (bez sazby
    z rejstriku): klic = (12 + 3 * (SIREN mod 97)) mod 97.
    """
    cislice = re.sub(r"\D", "", str(siren or ""))
    if len(cislice) != 9:
        return ""
    klic = (12 + 3 * (int(cislice) % 97)) % 97
    return "FR%02d%s" % (klic, cislice)


def _fr_adresa(sidlo):
    cast = " ".join(x for x in (
        sidlo.get("numero_voie"), sidlo.get("type_voie"), sidlo.get("libelle_voie")) if x)
    return cast or (sidlo.get("adresse") or "")


FR_POLE = ("siren", "nom_complet", "nom_raison_sociale", "sigle", "siege",
           "activite_principale", "nature_juridique", "date_creation", "etat_administratif")


def _fr_ocisti(data):
    def zmensi(r):
        r = {k: v for k, v in r.items() if k in FR_POLE}
        s = r.get("siege") or {}
        r["siege"] = {k: v for k, v in s.items() if k in (
            "numero_voie", "type_voie", "libelle_voie", "code_postal",
            "libelle_commune", "adresse", "etat_administratif")}
        return r
    return {"results": [zmensi(r) for r in data.get("results", [])]}


def fr_podle_nazvu(klient, nazev, pocet=15):
    url = INSEE_FR + "?" + urllib.parse.urlencode({"q": nazev, "per_page": min(pocet, 25)})
    data = json.loads(klient.ziskej(url, ocisti=_fr_ocisti))
    vysledky = []
    for r in data.get("results", []):
        sidlo = r.get("siege") or {}
        naf = r.get("activite_principale") or ""
        nace = taxonomie.naf_na_nace(naf)
        siren = r.get("siren") or ""
        aktivni = (sidlo.get("etat_administratif") or r.get("etat_administratif")) != "F"
        vysledky.append(Zaznam(
            jmeno=r.get("nom_complet") or r.get("nom_raison_sociale") or "",
            ulice=_fr_adresa(sidlo),
            psc=_psc(sidlo.get("code_postal")),
            mesto=sidlo.get("libelle_commune") or "",
            zeme="FR",
            dic=fr_dic_ze_siren(siren),
            reg_cislo=siren,
            reg_rejstrik="SIREN",
            nace=nace,
            nace_popis=taxonomie.nazev_nace(nace),
            nace_vse=naf,
            pravni_forma=r.get("nature_juridique") or "",
            datum_vzniku=r.get("date_creation") or "",
            aktivni=aktivni,
            zdroj="INSEE/INPI",
            odkaz="https://annuaire-entreprises.data.gouv.fr/entreprise/%s" % siren if siren else "",
            poznamka="zanikla firma (INSEE)" if not aktivni else "",
        ))
    return vysledky


# ---------------------------------------------------------------------------
# Singapur - ACRA (data.gov.sg, otevrena data)
# ---------------------------------------------------------------------------

SG_ACRA_POLE = ("uen", "entity_name", "entity_type_desc", "uen_status_desc",
               "reg_street_name", "reg_postal_code")


def _sg_ocisti(d):
    return {"result": {"records": [{k: v for k, v in r.items() if k in SG_ACRA_POLE}
                                   for r in d.get("result", {}).get("records", [])]}}


def _sg_na_zaznam(r):
    uen = r.get("uen") or ""
    aktivni = (r.get("uen_status_desc") or "").strip().lower() == "registered"
    return Zaznam(
        jmeno=r.get("entity_name") or "",
        ulice=r.get("reg_street_name") or "",
        psc=r.get("reg_postal_code") or "",
        mesto="Singapore" if r.get("reg_street_name") else "",
        zeme="SG",
        reg_cislo=uen,
        reg_rejstrik="UEN (ACRA)",
        pravni_forma=r.get("entity_type_desc") or "",
        aktivni=aktivni,
        zdroj="ACRA (data.gov.sg)",
        odkaz="https://www.uen.gov.sg/ueninternet/faces/pages/uenResults.jspx?uen=%s" % uen
              if uen else "",
        poznamka="stav v ACRA: %s" % r["uen_status_desc"]
                 if r.get("uen_status_desc") and not aktivni else "",
    )


def sg_podle_nazvu(klient, nazev, pocet=15):
    url = SG_ACRA + "?" + urllib.parse.urlencode({
        "resource_id": SG_ACRA_ZDROJ, "q": nazev, "limit": min(pocet, 30)})
    data = json.loads(klient.ziskej(url, ocisti=_sg_ocisti))
    return [_sg_na_zaznam(r) for r in data.get("result", {}).get("records", [])]


def sg_podle_uen(klient, uen):
    """Presny dotaz na jeden UEN - spolehlivejsi nez fulltextove hledani jmenem."""
    url = SG_ACRA + "?" + urllib.parse.urlencode({
        "resource_id": SG_ACRA_ZDROJ, "filters": json.dumps({"uen": uen}), "limit": 1})
    data = json.loads(klient.ziskej(url, ocisti=_sg_ocisti))
    zaznamy = data.get("result", {}).get("records", [])
    return _sg_na_zaznam(zaznamy[0]) if zaznamy else None


# ---------------------------------------------------------------------------
# Tchaj-wan - GCIS (data.gcis.nat.gov.tw, otevrena data)
# ---------------------------------------------------------------------------

def tw_podle_nazvu(klient, nazev, pocet=15):
    """
    GCIS bez filtru na stav vraci prazdno i pro bezne existujici firmy - proto
    je "Company_Status eq 01" (aktivni/schvalene zalozeni) soucasti dotazu,
    ne dodatecny filtr az na strane klienta.
    """
    url = TW_GCIS + "?" + urllib.parse.urlencode({
        "$format": "json",
        "$filter": "Company_Name like %s and Company_Status eq 01" % nazev,
        "$skip": 0, "$top": min(pocet, 30),
    })
    data = klient.ziskej(url, kontext=tw_ssl_kontext(), ocisti=lambda d: [
        {k: v for k, v in r.items() if k in (
            "Business_Accounting_NO", "Company_Name", "Company_Status_Desc",
            "Company_Location", "Company_Setup_Date")}
        for r in d] if isinstance(d, list) else d)
    zaznamy = json.loads(data)
    if not isinstance(zaznamy, list):
        return []
    vysledky = []
    for r in zaznamy:
        cislo = r.get("Business_Accounting_NO") or ""
        datum = r.get("Company_Setup_Date") or ""
        # tchajwanske datum je v minguo kalendari (rok - 1911), napr. "0760221"
        # = 1976-02-21 - pro cteni ve vystupu neni potreba prevadet, jen orezat
        vysledky.append(Zaznam(
            jmeno=r.get("Company_Name") or "",
            ulice=r.get("Company_Location") or "",
            zeme="TW",
            reg_cislo=cislo,
            reg_rejstrik="統一編號 (GCIS)",
            datum_vzniku=datum,
            zdroj="GCIS",
            odkaz="https://data.gcis.nat.gov.tw/od/detail?oid=6BBA2268-1367-4B42-9CCA-BC17499EBE8C",
            poznamka="",
        ))
    return vysledky


# ---------------------------------------------------------------------------
# Wikidata - zalozni zdroj pro velke zahranicni firmy
# ---------------------------------------------------------------------------

WIKIDATA_POLE = ("P452", "P17", "P159", "P3608", "P1278", "P297")


def _wikidata_ocisti(data):
    """
    Odpovedi wbgetentities jsou obrovske (entita zeme ma tisice tvrzeni).
    Do kese ukladame jen popisky a tvrzeni, ktera cteme.
    """
    entity = {}
    for qid, e in (data.get("entities") or {}).items():
        entity[qid] = {
            "labels": {j: v for j, v in (e.get("labels") or {}).items() if j in ("en", "cs", "de")},
            "claims": {p: v for p, v in (e.get("claims") or {}).items() if p in WIKIDATA_POLE},
        }
    return {"entities": entity}


def wikidata_podle_nazvu(klient, nazev, pocet=5, i_kratky_nazev=False):
    """
    `i_kratky_nazev=True` navic zkusi hledani pod zkracenym nazvem (bez
    pravni formy) a vysledky sloucí, i kdyz plny nazev neco naslo. Rejstriky
    a VIES vraceji plny pravni nazev ("ABB Schweiz AG", "ENI SPA"), Wikidata
    ale firmy vetsinou vede pod kratkym nazvem clanku ("ABB", "Eni") -
    hledani s plnym nazvem tak casto neni prazdne, jen vrati nekoho jineho
    (dceřinou spolecnost, pobocku).

    Vyplati se to jen pri doplnovani oboru k uz jednoznacne identifikovane
    firme (vic kandidatu vadit nemuze - obor se bere z prvniho s dost
    podobnym jmenem). U hledani samotne identity firmy by to naopak skodilo:
    kratky nazev typicky vrati vic stejne se jmenujicich firem/dceřinych
    spolecnosti se stejnym skore, coz zvysuje falesnou "VICE_SHOD"
    nejednoznacnost a muze prebit spravneho kandidata horsim.
    """
    def hledej(text):
        url = WIKIDATA_API + "?" + urllib.parse.urlencode({
            "action": "wbsearchentities", "search": text, "language": "en",
            "uselang": "en", "type": "item", "limit": pocet, "format": "json"})
        return json.loads(klient.ziskej(url, ocisti=lambda d: {"search": [
            {"id": h.get("id")} for h in d.get("search", [])]})).get("search", [])

    hledani = hledej(nazev)
    jadro = jadro_pro_hledani(re.sub(r"[,.]+$", "", nazev.replace(",", " ")))
    jadro = re.sub(r"[\s,.\-]+$", "", jadro)
    ma_smysl_zkusit_jadro = jadro.lower() != nazev.lower() and jadro
    if ma_smysl_zkusit_jadro and (not hledani or i_kratky_nazev):
        videne = {h["id"] for h in hledani}
        hledani += [h for h in hledej(jadro) if h["id"] not in videne]
    if not hledani:
        return []

    qidy = [h["id"] for h in hledani[:8]]
    url = WIKIDATA_API + "?" + urllib.parse.urlencode({
        "action": "wbgetentities", "ids": "|".join(qidy),
        "props": "claims|labels", "languages": "en|cs|de", "format": "json"})
    entity = json.loads(klient.ziskej(url, ocisti=_wikidata_ocisti)).get("entities", {})

    def tvrzeni(claims, prop):
        vysledek = []
        for s in claims.get(prop, []):
            dv = s.get("mainsnak", {}).get("datavalue", {}).get("value")
            if isinstance(dv, dict) and "id" in dv:
                vysledek.append(dv["id"])
            elif isinstance(dv, str):
                vysledek.append(dv)
        return vysledek

    # QID odvetvi / zeme / sidla je treba prelozit na text
    odkazovane = set()
    for qid in qidy:
        c = entity.get(qid, {}).get("claims", {})
        odkazovane.update(tvrzeni(c, "P452")[:5] + tvrzeni(c, "P17")[:1] + tvrzeni(c, "P159")[:1])
    popisky, iso = {}, {}
    if odkazovane:
        url = WIKIDATA_API + "?" + urllib.parse.urlencode({
            "action": "wbgetentities", "ids": "|".join(sorted(odkazovane)[:50]),
            "props": "labels|claims", "languages": "en|cs", "format": "json"})
        for qid, e in json.loads(
                klient.ziskej(url, ocisti=_wikidata_ocisti)).get("entities", {}).items():
            lab = e.get("labels", {})
            popisky[qid] = (lab.get("en") or lab.get("cs") or {}).get("value", "")
            kody = [s.get("mainsnak", {}).get("datavalue", {}).get("value")
                    for s in e.get("claims", {}).get("P297", [])]
            if kody and isinstance(kody[0], str):
                iso[qid] = kody[0]

    vysledky = []
    for qid in qidy:
        e = entity.get(qid, {})
        c = e.get("claims", {})
        obory_qid = tvrzeni(c, "P452")[:5]
        obory = [popisky.get(q, "") for q in obory_qid]
        zeme_q = tvrzeni(c, "P17")[:1]
        sidlo_q = tvrzeni(c, "P159")[:1]
        dic = next((v for v in tvrzeni(c, "P3608") if isinstance(v, str)), "")
        lei = next((v for v in tvrzeni(c, "P1278") if isinstance(v, str)), "")
        lab = e.get("labels", {})
        vysledky.append(Zaznam(
            jmeno=(lab.get("en") or lab.get("cs") or lab.get("de") or {}).get("value", ""),
            mesto=popisky.get(sidlo_q[0], "") if sidlo_q else "",
            zeme=iso.get(zeme_q[0], "") if zeme_q else "",
            dic=dic, lei=lei,
            obory=obory_qid,
            zdroj="Wikidata",
            odkaz="https://www.wikidata.org/wiki/%s" % qid,
            poznamka="obor dle Wikidata: %s" % ", ".join(o for o in obory if o) if any(obory) else "",
        ))
    return vysledky


# ---------------------------------------------------------------------------
# VIES - overeni DIC v EU
# ---------------------------------------------------------------------------

def rozloz_dic(dic):
    if not dic:
        return None, None
    d = re.sub(r"[^A-Za-z0-9]", "", str(dic)).upper()
    m = re.match(r"^([A-Z]{2})(\w+)$", d)
    return (m.group(1), m.group(2)) if m else (None, None)


# VIES vraci docasne vypadky sluzby jako HTTP 200 s isValid:false a timto
# priznakem v "userError" - ne jako chybu. Bez rozliseni by to vypadalo
# jako platne overeni "DIC neexistuje", pritom sluzba jen byla docasne
# preteizena (typicky pri vic soubeznych dotazech na stejny stat).
VIES_DOCASNE_CHYBY = {"MS_MAX_CONCURRENT_REQ", "MS_UNAVAILABLE", "GLOBAL_MAX_CONCURRENT_REQ",
                      "SERVICE_UNAVAILABLE", "TIMEOUT", "SERVER_BUSY"}


def vies_over(klient, dic, pokusy=5):
    cc, num = rozloz_dic(dic)
    if not cc or cc not in EU_STATY:
        return None
    url = VIES_API.format(cc=cc, num=num)
    for pokus in range(1, pokusy + 1):
        data = json.loads(klient.ziskej(url))
        chyba = data.get("userError")
        if chyba in VIES_DOCASNE_CHYBY:
            klient.zapomen(url)   # nezustane v kesi jako by to byla platna odpoved
            if pokus < pokusy:
                time.sleep(min(2 ** pokus, 6))
                continue
            raise RuntimeError("docasne nedostupne (%s)" % chyba)
        return {"platne": bool(data.get("isValid")),
                "jmeno": (data.get("name") or "").strip(" -"),
                "adresa": (data.get("address") or "").strip(" -")}


# ---------------------------------------------------------------------------
# Odkaz na rucni dohledani - zeme bez napojeneho rejstriku ani u GLEIF/Wikidat
# ---------------------------------------------------------------------------

# Domena narodniho rejstriku pro zeme, kde nemame API (viz README - overeno,
# ze bezklicove hromadne API neexistuje). Pouziva se jen k sestaveni odkazu
# na vyhledavani "site:domena nazev" - clovek pak dotaz jen otevre, nic se
# tim neautomatizuje ani neobchazi.
ZEME_MANUALNI_REJSTRIK = {
    "DE": "handelsregister.de",
    "NL": "kvk.nl",
    "AT": "justiz.gv.at",
    "BE": "kbopub.economie.fgov.be",
    "CH": "zefix.ch",
    "IT": "registroimprese.it",
    "ES": "sede.registradores.org",
    "HU": "e-cegjegyzek.hu",
    "PL": "krs-online.com.pl",
    "RO": "onrc.ro",
    "BG": "portal.registryagency.bg",
    "TR": "ticaretsicil.gov.tr",
    "MY": "ssm-einfo.my",
    "HK": "cr.gov.hk",
    "CA": "corporationscanada.ic.gc.ca",
    "GB": "find-and-update.company-information.service.gov.uk",
}

# Zeme pripojene k BRIS (European Business Registers Interconnection System) -
# u tech je alternativou i centralni EU vyhledavani (rucne, viz README)
BRIS_ZEME = {"AT", "BE", "BG", "CY", "CZ", "DE", "DK", "EE", "EL", "ES", "FI",
             "FR", "HR", "HU", "IE", "IT", "LT", "LU", "LV", "MT", "NL", "PL",
             "PT", "RO", "SE", "SI", "SK", "IS", "LI", "NO"}


def manualni_odkaz(zeme, nazev):
    """
    Kdyz se firma nenajde a zeme nema napojeny zadny rejstrik, vrati odkaz
    na predpripraveny vyhledavaci dotaz (Google omezeny na spravnou domenu,
    pripadne centralni BRIS) - usetri rucni psani nazvu do vyhledavace.
    """
    if not nazev:
        return ""
    domena = ZEME_MANUALNI_REJSTRIK.get(zeme)
    if domena:
        dotaz = 'site:%s "%s"' % (domena, nazev)
        return "https://www.google.com/search?q=" + urllib.parse.quote(dotaz)
    if zeme in BRIS_ZEME:
        return "https://e-justice.europa.eu/content_find_a_company-489-en.do"
    return ""


# ---------------------------------------------------------------------------
# Zpracovani jednoho radku
# ---------------------------------------------------------------------------

def _normalizuj_mesto(s):
    return re.sub(r"[^a-z0-9]+", " ", bez_diakritiky(str(s or "")).lower()).strip()


def _normalizuj_psc(s):
    return re.sub(r"\D", "", str(s or ""))


def _shoduje_se_adresa(hledana, k):
    """
    Porovna zadanou adresu s adresou kandidata. Vraci (bonus, duvod|None).
    Shoda mesta nebo PSC je silny signal, ze jde o spravny subjekt (obzvlast
    u firem se skupinou stejne se jmenujicich dcerinych spolecnosti v ruznych
    mestech) - naopak jasny nesoulad (zadano konkretni mesto, kandidat sidli
    jinde) je duvod k ostrazitosti podobne jako spatna zeme.
    """
    if not hledana or not any(hledana.values()):
        return 0.0, None
    mesto_h, psc_h = _normalizuj_mesto(hledana.get("mesto")), _normalizuj_psc(hledana.get("psc"))
    mesto_k, psc_k = _normalizuj_mesto(k.mesto), _normalizuj_psc(k.psc)

    if psc_h and psc_k:
        if psc_h == psc_k:
            return 0.06, None
        if psc_h[:2] != psc_k[:2]:
            return -0.05, "PSC %s misto %s" % (k.psc, hledana.get("psc"))
    if mesto_h and mesto_k:
        if mesto_h == mesto_k or mesto_h in mesto_k or mesto_k in mesto_h:
            return 0.06, None
        return -0.05, "sidlo v %s misto %s" % (k.mesto, hledana.get("mesto"))
    return 0.0, None


def skore_kandidata(nazev, zeme, k, adresa=None):
    """
    Slozene skore jednoho kandidata: podobnost nazvu plus prirazky za to,
    jak dobre zaznam sedi na zbytek zadani.

    Diky prirazkam se da rozhodnout i tam, kde je nazev stejne podobny u vic
    firem - typicky "Danone S.A." sedi na DANONE (FR) i DANONE PORTUGAL (PT)
    a rozhodne az zadana zeme (nebo adresa, je-li k dispozici).

    Vrati (celkove skore, skore nazvu, duvody prirazek).
    """
    jmena = [j for j in (k.jmena or [k.jmeno]) if j]
    if not jmena:
        return 0.0, 0.0, []
    skore = max(skore_shody(nazev, j) for j in jmena)
    syrove = max(skore_syrove(nazev, j) for j in jmena)

    bonus, duvody = 0.0, []
    if zeme and k.zeme:
        if k.zeme == zeme:
            bonus += 0.08
        else:
            bonus -= 0.12
            duvody.append("sidlo v %s misto %s" % (k.zeme, zeme))
    adr_bonus, adr_duvod = _shoduje_se_adresa(adresa, k)
    bonus += adr_bonus
    if adr_duvod:
        duvody.append(adr_duvod)
    if not k.aktivni:
        bonus -= 0.15
        duvody.append("neaktivni subjekt")
    if k.nace:
        bonus += 0.03
    if k.reg_cislo or k.ico:
        bonus += 0.02
    if k.dic:
        bonus += 0.01
    if k.obory:
        bonus += 0.02
    # syrova podobnost rozhodne remizy, kde normalizace zahodila prave to
    # slovo, kterym se kandidati lisi ("Group", "Holding")
    return round(skore + bonus + 0.05 * syrove, 4), skore, duvody


def vyber_nejlepsi(kandidati, nazev, prah_ok, prah_overit, zeme=None, adresa=None):
    """
    Vrati (nejlepsi, stav, prehled kandidatu).

    Pri vice srovnatelnych shodach se vzdy vybere jeden zaznam - ten
    s nejvyssim slozenym skore - a stav je VYBRANO. Rucni dohledavani
    v seznamu o stovkach dodavatelu se tim nahradi kontrolou par radku;
    ostatni kandidati zustavaji vypsani v poznamce.
    """
    if not kandidati:
        return None, STAV_NENALEZENO, []
    ohodnocene = sorted(((skore_kandidata(nazev, zeme, k, adresa), k) for k in kandidati),
                        key=lambda x: -x[0][0])
    (celkem, skore, duvody), nejlepsi = ohodnocene[0]
    nejlepsi.shoda = "%.0f%%" % (skore * 100)

    prehled = ["%s [%s%s] %.0f%%" % (k.jmeno, k.zeme + " " if k.zeme else "",
                                     k.reg_cislo or k.ico or k.lei or k.mesto or "?", s * 100)
               for (c, s, _), k in ohodnocene[:5] if s > 0.3]

    if skore >= prah_ok:
        # OSVC: kdyz mame vic stejnojmennych osob (aspon jedna dalsi se stejnym
        # jmenem mezi VSEMI kandidaty, ne jen mezi temi s vysokym skore) a
        # zadana adresa nesedi ani na tu nejlepe skorujici, nejde jen o "trochu
        # nejiste" - presna kombinace jmeno+adresa mezi kandidaty vubec neni.
        # Radeji NENALEZENO, nez tipovat (byt s upozornenim v poznamce), ktery
        # z nekolika stejnojmennych je ten spravny - propsani cizi adresy/ICO
        # by bylo horsi nez prazdno. Musi se resit pred VYBRANO/OVERIT - jinak
        # "srovnatelni" (pocitane jen z podobnosti jmena) skoro vzdy odchyti
        # tenhle pripad driv a schova ho za VYBRANO.
        if nejlepsi.pravni_forma in OSVC_PRAVNI_FORMY and adresa and any(adresa.values()):
            adresa_nesedi = any(d.startswith(("sidlo v", "PSC ")) for d in duvody)
            stejne_jmeno = sum(
                1 for k in kandidati
                if k.pravni_forma in OSVC_PRAVNI_FORMY
                and normalizuj_nazev(k.jmeno) == normalizuj_nazev(nejlepsi.jmeno))
            if adresa_nesedi and stejne_jmeno > 1:
                return None, STAV_NENALEZENO, prehled

        # kolik dalsich kandidatu je jmenem stejne dobrych
        srovnatelni = sum(1 for (c, s, _), _ in ohodnocene[1:] if s >= prah_ok)
        if srovnatelni:
            druhy_celkem = ohodnocene[1][0][0]
            nejlepsi.poznamka = "; ".join(p for p in (
                nejlepsi.poznamka,
                "vybrano z %d srovnatelnych shod, druhy v poradi o %.2f bodu niz"
                % (srovnatelni + 1, celkem - druhy_celkem),
                ", ".join(duvody),
            ) if p)
            return nejlepsi, STAV_VYBRANO, prehled
        if duvody:
            nejlepsi.poznamka = "; ".join(p for p in (nejlepsi.poznamka,
                                                      "pozor: " + ", ".join(duvody)) if p)
            return nejlepsi, STAV_OVERIT, prehled
        # OSVC nalezena jen podle jmena osoby, bez adresy k rozliseni - jmeno
        # samo o sobe neni jednoznacne (na rozdil od nazvu firmy), takze i
        # jediny takto ziskany kandidat neni jisty (viz OSVC_PRAVNI_FORMY) -
        # jina stejnojmenna osoba mohla byt jen mimo nactenych "pocet" kandidatu.
        if nejlepsi.pravni_forma in OSVC_PRAVNI_FORMY and not (adresa and any(adresa.values())):
            nejlepsi.poznamka = "; ".join(p for p in (
                nejlepsi.poznamka,
                "pozor: nalezeno jen podle jmena osoby (OSVC) bez adresy k rozliseni - "
                "u beznych jmen muze v ARES existovat vic stejnojmennych osob, "
                "doplnte adresu nebo ICO pro jistotu",
            ) if p)
            return nejlepsi, STAV_OVERIT, prehled
        return nejlepsi, STAV_OK, prehled
    if skore >= prah_overit:
        return nejlepsi, STAV_OVERIT, prehled
    return nejlepsi, STAV_NENALEZENO, prehled


def zpracuj_radek(vstup, klient, n):
    nazev = (vstup.get("nazev") or "").strip()
    ico = (vstup.get("ico") or "").strip()
    dic = (vstup.get("dic") or "").strip()
    zeme = (vstup.get("zeme") or "").strip().upper()[:2]
    hledana_adresa = {
        "ulice": (vstup.get("ulice") or "").strip(),
        "psc": (vstup.get("psc") or "").strip(),
        "mesto": (vstup.get("mesto") or "").strip(),
    }

    z = Zaznam(hledany_nazev=nazev or ico or dic)
    poznamky = []

    try:
        if not ico and dic:
            cc, num = rozloz_dic(dic)
            if cc == "CZ" and num and num.isdigit():
                ico = num
        if not ico and nazev and re.fullmatch(r"\d{6,8}", re.sub(r"\s", "", nazev)):
            ico = re.sub(r"\s", "", nazev)

        # 1) presna identifikace podle ICO
        if ico and zeme in ("", "CZ") and not n["bez_ares"]:
            try:
                z = ares_podle_ica(klient, ico)
                z.hledany_nazev = nazev or ico
                z.shoda = "100%" if not nazev else "%.0f%%" % (skore_shody(nazev, z.jmeno) * 100)
                z.stav = STAV_OK
                # Presne cislo (ICO, nebo ICO odvozene z DIC) muze byt na
                # vstupu spatne - preklep, spatne zkopirovany radek apod.
                # U OSVC muze byt jinak stejne jmeno zadano se spravnou
                # adresou, ale s cislem patricim jinemu stejnojmennemu
                # clovku - bez kontroly adresy by se to tise vzalo jako
                # jiste OK. Kontroluje se JEN u OSVC (fyzickych osob) -
                # u firem je bezne, ze provozni/korespondencni adresa na
                # vstupu neodpovida formalnimu sidlu v ARES, a takovy
                # nesoulad nic neznamena (na rozdil od skutecne jine
                # osoby/firmy u shodneho jmena).
                if z.pravni_forma in OSVC_PRAVNI_FORMY:
                    adr_bonus, adr_duvod = _shoduje_se_adresa(hledana_adresa, z)
                    if adr_duvod and adr_bonus < 0:
                        z.stav = STAV_OVERIT
                        poznamky.append(
                            "pozor: zadane ICO/DIC bylo v ARES nalezeno, ale zadana "
                            "adresa neodpovida (%s) - zkontrolujte, jestli cislo "
                            "patri ke spravne osobe" % adr_duvod)
            except Exception as e:
                poznamky.append("ARES podle ICO: %s" % e)
                # subjekt jiz neni v hlavnim indexu ARES - zkusit Verejny
                # rejstrik, ktery drzi historii i po vymazu (datum a duvod)
                try:
                    vymaz = ares_vr_vymaz(klient, ico)
                    if vymaz:
                        poznamky.append("vymazan z rejstriku %s%s" % (
                            vymaz["datum"],
                            (", duvod: %s" % vymaz["duvod"]) if vymaz["duvod"] else ""))
                except Exception:
                    pass

        # 1b) presna identifikace zahranicniho subjektu podle zadaneho cisla
        # (LEI, narodni registracni cislo typu HRB/SIREN/CIK zapsane
        # do sloupce ICO) - spolehlivejsi nez hledani jmenem, protoze cislo
        # je na rozdil od nazvu jednoznacne
        if z.stav != STAV_OK and ico and zeme and zeme != "CZ":
            try:
                kandidati = []
                if zeme == "FR" and not n["bez_fr"]:
                    cislice = re.sub(r"\D", "", ico)
                    kandidati = [k for k in fr_podle_nazvu(klient, ico, 10)
                                if k.reg_cislo == cislice]
                elif zeme == "US" and not n["bez_edgar"]:
                    nalezeny = edgar_podle_cik(klient, ico)
                    kandidati = [nalezeny] if nalezeny else []
                elif zeme == "SG" and not n["bez_sg"]:
                    nalezeny = sg_podle_uen(klient, ico.upper())
                    kandidati = [nalezeny] if nalezeny else []
                elif zeme == "DE" and not n["bez_de"] and os.path.exists(DE_REGISTER_DB):
                    druh, cislo = de_rozloz_reg_cislo(ico)
                    if druh:
                        kandidati = de_podle_registru(druh, cislo)
                elif zeme == "GB" and not n["bez_gb"] and os.path.exists(GB_REGISTER_DB):
                    nalezeny = gb_podle_cisla(ico)
                    kandidati = [nalezeny] if nalezeny else []
                if not kandidati and not n["bez_gleif"]:
                    if re.fullmatch(r"[A-Za-z0-9]{20}", ico):
                        nalezeny = gleif_podle_lei(klient, ico.upper())
                        kandidati = [nalezeny] if nalezeny else []
                    else:
                        kandidati = gleif_podle_registrovane(klient, ico, zeme)

                if len(kandidati) == 1:
                    z = kandidati[0]
                    z.hledany_nazev = nazev or ico
                    z.stav = STAV_OK
                    z.shoda = "100%"
                    z.poznamka = "; ".join(p for p in (
                        z.poznamka, "nalezeno podle zadaneho identifikacniho cisla") if p)
                elif len(kandidati) > 1:
                    # vzacny pripad - stejne cislo u vice zaznamu (napr.
                    # znovupouzite HRB po zaniku puvodni firmy). Bez nazvu
                    # k rozhodnuti aspon vypsat kandidaty k rucni kontrole
                    if nazev:
                        nej, stav_id, prehled_id = vyber_nejlepsi(
                            kandidati, nazev, n["prah_ok"], n["prah_overit"], zeme,
                            hledana_adresa)
                        if nej is not None and stav_id != STAV_NENALEZENO:
                            nej.hledany_nazev = nazev
                            nej.kandidati = prehled_id
                            z = nej
                    else:
                        z.kandidati = ["%s [%s %s]" % (k.jmeno, k.zeme, k.reg_cislo)
                                      for k in kandidati[:5]]
                        poznamky.append(
                            "zadane cislo odpovida %d ruznym zaznamum - chybi nazev "
                            "k rozliseni" % len(kandidati))
            except Exception as e:
                poznamky.append("presna identifikace podle cisla: %s" % e)

        # 1c) presna identifikace podle DIC pres VIES - kdyz ICO cestu
        # nevyresilo. Data jsou chudsi (jen jmeno a adresa, zadne NACE ani
        # registracni cislo), ale DIC je - na rozdil od nazvu - jednoznacne,
        # takze se zaradi jeste pred hledanim jmenem. Funguje jen pro EU DIC
        # (vies_over pro ostatni vrati None).
        if z.stav != STAV_OK and dic:
            try:
                v = vies_over(klient, dic)
                if v and v["platne"] and v["jmeno"]:
                    cc, _ = rozloz_dic(dic)
                    z.jmeno = v["jmeno"]
                    z.ulice = v["adresa"]
                    z.zeme = z.zeme or cc or ""
                    z.dic = dic
                    z.dic_overeno = "ANO"
                    z.stav = STAV_OK
                    z.shoda = "100%"
                    z.zdroj = "VIES"
                    z.hledany_nazev = nazev or dic
                    poznamky.append("nalezeno podle DIC pres VIES")
            except Exception as e:
                poznamky.append("VIES (hledani podle DIC): %s" % e)

        # 2) hledani podle nazvu
        if z.stav != STAV_OK and nazev:
            nejlepsi, stav, prehled, nej_skore = None, STAV_NENALEZENO, [], -1.0
            # poradi stavu pri rozhodovani, ktery zdroj vyhraje
            vaha = {STAV_OK: 3, STAV_VYBRANO: 2, STAV_OVERIT: 1, STAV_NENALEZENO: 0}

            def zkus(funkce, *args):
                nonlocal nejlepsi, stav, prehled, nej_skore
                if stav == STAV_OK:            # jednoznacna shoda, dal nehledame
                    return
                try:
                    k_nejlepsi, k_stav, k_prehled = vyber_nejlepsi(
                        funkce(klient, *args), nazev, n["prah_ok"], n["prah_overit"], zeme,
                        hledana_adresa)
                except Exception as e:
                    poznamky.append("%s: %s" % (funkce.__name__, e))
                    return
                if k_nejlepsi is None:
                    return
                k_skore = float(k_nejlepsi.shoda.rstrip("%") or 0) / 100.0
                if (vaha[k_stav], k_skore) > (vaha[stav], nej_skore):
                    nejlepsi, stav, nej_skore = k_nejlepsi, k_stav, k_skore
                    prehled = k_prehled or prehled

            # poradi: nejdriv narodni rejstriky, ktere nesou i obor cinnosti,
            # az pak celosvetove zdroje bez oboru (GLEIF) a bez presneho NACE
            # (Wikidata)
            if zeme in ("", "CZ") and not n["bez_ares"]:
                zkus(ares_podle_nazvu, nazev, n["pocet"])
            if zeme in ("", "SK") and not n["bez_sk"]:
                zkus(rpo_sk_podle_nazvu, nazev, min(n["pocet"], 20))
            if zeme == "FR" and not n["bez_fr"]:
                zkus(fr_podle_nazvu, nazev, n["pocet"])
            if zeme == "SG" and not n["bez_sg"]:
                zkus(sg_podle_nazvu, nazev, n["pocet"])
            if zeme == "TW" and not n["bez_tw"]:
                zkus(tw_podle_nazvu, nazev, n["pocet"])
            if zeme in ("", "US") and not n["bez_edgar"]:
                zkus(edgar_podle_nazvu, nazev, n["pocet"])
            if zeme == "DE" and n["openregister_klic"]:
                zkus(openregister_podle_nazvu, nazev, n["openregister_klic"], n["pocet"])
            elif zeme == "DE" and not n["bez_de"] and os.path.exists(DE_REGISTER_DB):
                zkus(de_podle_nazvu, nazev, n["pocet"])
            if zeme == "GB" and not n["bez_gb"] and os.path.exists(GB_REGISTER_DB):
                zkus(gb_podle_nazvu, nazev, n["pocet"])
            if zeme in SCORIS_ZEME and n["scoris_klic"]:
                zkus(scoris_podle_nazvu, nazev, n["scoris_klic"], zeme, n["pocet"])
            if zeme != "CZ" and not n["bez_gleif"]:
                zkus(gleif_podle_nazvu, nazev, zeme or None, n["pocet"])
            if zeme != "CZ" and not n["bez_wikidata"]:
                zkus(wikidata_podle_nazvu, nazev, 5)

            if nejlepsi is not None and stav != STAV_NENALEZENO:
                if nejlepsi.poznamka:
                    poznamky.insert(0, nejlepsi.poznamka)
                nejlepsi.hledany_nazev = nazev
                nejlepsi.stav = stav
                nejlepsi.kandidati = prehled
                z = nejlepsi
            else:
                # nic dost podobneho - data kandidatu se do vystupu neprenaseji,
                # jen se nabidnou v poznamce k rucni kontrole
                z.hledany_nazev = nazev
                z.stav = STAV_NENALEZENO
                z.kandidati = prehled

        if not nazev and not ico and not dic:
            z.stav = STAV_CHYBA
            poznamky.append("prazdny radek vstupu")

        z.ico = z.ico or ico
        z.dic = z.dic or dic
        z.zeme = z.zeme or zeme

        # 3) upresneni oboru cinnosti
        if z.stav != STAV_NENALEZENO and not n["bez_ares"]:
            doplr_prevazujici_nace(klient, z)
            # RES uvadi jako prevazujici cinnost obecny/podpurny kod (napr.
            # pronajem nemovitosti, 6820) - u casti firem jde skutecne o hlavni
            # predmet podnikani (typicky cista majetkova/holdingova entita
            # v ramci skupiny - pak je to spravny udaj, ne chyba), u jinych
            # jde spis o formalni registraci "pro jistotu" pri zalozeni firmy.
            # Nejde rozeznat, ktera situace nastala, bez skutecne znalosti
            # firmy - zaznam se proto NEPREKLADA jinam (bylo by to hadani,
            # napr. z nazvu firmy - viz README), jen se oznaci k rucni/LLM
            # kontrole (--export-nezarazene zahrne i tyto zaznamy, ne jen
            # XXX-00, viz nace_nejisty).
            zakladni_kod = re.sub(r"\D", "", str(z.nace or ""))
            if zakladni_kod and zakladni_kod in PODPURNE_NACE:
                z.nace_nejisty = True
                neurceno = zakladni_kod.strip("0") == ""   # "00"/"0000" - doslova "neurceno"
                popis_kodu = ("nema v ARES zapsanou zadnou hlavni cinnost (kod %s = neurceno)"
                             % z.nace if neurceno else
                             "ma jako hlavni cinnost obecny/podpurny kod %s (napr. "
                             "pronajem/spravni cinnosti)" % z.nace)
                if len(set(z.nace_vse.split(","))) > 1:
                    poznamky.append(
                        "firma %s - zkontrolujte NACE (vsechny), skutecny obor muze "
                        "byt jiny" % popis_kodu)
                else:
                    duvod = ("firma obor jeste zrejme nenahlasila" if neurceno else
                             "muze jit o majetkovou/holdingovou firmu se spravnym "
                             "udajem, nebo jen o formalni registraci")
                    poznamky.append(
                        "firma %s (jediny zapsany kod) - %s; doporucena kontrola "
                        "pres --export-nezarazene" % (popis_kodu, duvod))
                    if z.stav == STAV_OK:
                        z.stav = STAV_OVERIT
        if z.stav != STAV_NENALEZENO and not n["bez_sk"]:
            doplnit_sk_nace(klient, z)
        if z.stav != STAV_NENALEZENO and n["openregister_klic"]:
            try:
                doplnit_openregister_nace(klient, z, n["openregister_klic"])
            except Exception as e:
                poznamky.append("OpenRegister.de: %s" % e)
        if z.stav != STAV_NENALEZENO and n["scoris_klic"]:
            try:
                doplnit_scoris_detail(klient, z, n["scoris_klic"])
            except Exception as e:
                poznamky.append("Scoris: %s" % e)
        if (z.stav not in (STAV_NENALEZENO, STAV_CHYBA) and not z.nace and not z.obory
                and z.zdroj != "Wikidata" and not n["bez_wikidata"]):
            try:
                shodne = [w for w in wikidata_podle_nazvu(klient, z.jmeno or nazev, 5,
                                                          i_kratky_nazev=True)
                         if skore_shody(z.jmeno or nazev, w.jmeno) >= n["prah_ok"]]
                # stejny nazev v ruznych jazykovych mutacich casto vede na vic
                # QID - obor cinnosti byva zapsany jen u jednoho z nich
                w = next((c for c in shodne if c.obory), shodne[0] if shodne else None)
                if w is not None:
                    if w.poznamka:
                        poznamky.append(w.poznamka)
                    z.obory = w.obory
                    if not z.dic and w.dic:
                        z.dic = w.dic
                    if not z.lei and w.lei:
                        z.lei = w.lei
            except Exception as e:
                poznamky.append("Wikidata: %s" % e)

        # 4) overeni DIC pres VIES
        if n["vies"] and not z.dic and z.zeme == "CZ" and z.ico and "DPH" in z.poznamka:
            try:
                v = vies_over(klient, "CZ" + z.ico)
                if v and v["platne"]:
                    z.dic = "CZ" + z.ico
                    z.dic_overeno = "ANO"
                    poznamky.append("DIC doplneno a overeno pres VIES")
            except Exception as e:
                poznamky.append("VIES (odvozeni DIC): %s" % e)

        if n["vies"] and z.dic and not z.dic_overeno:
            try:
                v = vies_over(klient, z.dic)
                if v is not None:
                    z.dic_overeno = "ANO" if v["platne"] else "NE"
                    if v["platne"]:
                        if not z.jmeno and v["jmeno"]:
                            z.jmeno = v["jmeno"]
                        if not z.ulice and v["adresa"]:
                            z.ulice = v["adresa"]
            except Exception as e:
                poznamky.append("VIES: %s" % e)

        # 5) narodni registracni cislo -> citelny nazev rejstriku / pravni forma
        if z.zdroj == "GLEIF" and not n["bez_gleif_popisy"]:
            try:
                if z.reg_rejstrik:
                    z.reg_rejstrik = gleif_popis_rejstriku(klient, z.reg_rejstrik)
                if z.pravni_forma and len(z.pravni_forma) == 4:
                    z.pravni_forma = gleif_popis_formy(klient, z.pravni_forma)
            except Exception:
                pass

        # 7) vlastni taxonomie - jen z fakticke podkladu: NACE z rejstriku ma
        # prednost, pak obor z Wikidat (strukturovany QID, ne odhad z textu).
        # Bez ani jednoho jde zaznam do XXX-00 - kategorie se z nazvu firmy
        # nehada (hrozi false positive, viz README)
        k = taxonomie.zarad(nace=z.nace, mapa=n["mapa"], kategorie=n["kategorie_ciselnik"],
                            obory=z.obory, mapa_oboru=n["mapa_oboru"])
        z.kod_kategorie = k["kod"]
        z.kategorie = k["kategorie"]
        z.skupina = k["skupina"]
        z.zdroj_kategorie = k["zdroj"]
        if not z.nace and k["nace"]:
            # NACE se u zahranicnich firem bez rejstriku pocita jen jako
            # odhad z oboru cinnosti - proto se oznaci zvlast, ne jako
            # oficialni udaj z rejstriku
            z.nace = k["nace"]
            z.nace_zdroj = "odhad z oboru (Wikidata)"
        if not z.nace_popis:
            z.nace_popis = taxonomie.nazev_nace(z.nace)
        if z.nace and not z.nace_zdroj:
            z.nace_zdroj = z.zdroj

        # 8) je zaznam za aktivni subjekt?
        if not z.aktivni and z.stav not in (STAV_NENALEZENO, STAV_CHYBA):
            poznamky.append("subjekt neni v rejstriku veden jako aktivni")

    except Exception as e:
        z.stav = STAV_CHYBA
        poznamky.append("chyba: %r" % e)

    if z.kandidati and z.stav != STAV_OK:
        poznamky.append("kandidati: " + " | ".join(z.kandidati))
    z.poznamka = "; ".join(p for p in poznamky if p)

    # zaznam bez shody se do datovych sloupcu nepropisuje (viz README), ale
    # hledany nazev musi zustat viditelny i v zakladnim sloupci Jmeno -
    # jinak firma z nenalezene shody v exportu "zmizi"
    if not z.jmeno:
        z.jmeno = z.hledany_nazev or nazev or ico or dic

    # u nenalezene firmy ze zeme bez napojeneho rejstriku aspon predpripravit
    # odkaz na rucni vyhledani - usetri psani nazvu do vyhledavace
    if z.stav == STAV_NENALEZENO and not z.odkaz:
        z.odkaz = manualni_odkaz(z.zeme, z.jmeno)

    # nejlepsi cislo k rucnimu vlozeni do sloupce ICO pri druhem behu
    # (viz rezim --jen-id) - ICO/registracni cislo je citelnejsi nez LEI,
    # proto ma prednost
    z.identifikator = z.ico or z.reg_cislo or z.lei

    return z


# ---------------------------------------------------------------------------
# Vstup / vystup
# ---------------------------------------------------------------------------

MAPOVANI_SLOUPCU = {
    "nazev": ("nazev", "název", "jmeno", "jméno", "name", "firma", "company", "company name",
              "nazev spolecnosti", "název společnosti", "dodavatel", "supplier",
              "supplier name", "vendor", "vendor name", "obchodni jmeno", "obchodní jméno",
              "lieferant", "firmenname"),
    "ico": ("ico", "ičo", "ic", "reg no", "registration number", "company id", "identifikacni cislo"),
    "dic": ("dic", "dič", "vat", "vat id", "vat number", "ust-idnr", "ustidnr", "tax id",
            "st.-nr.", "st-nr"),
    "zeme": ("zeme", "země", "country", "stat", "stát", "land", "iso", "kod zeme"),
    "ulice": ("ulice", "street", "address", "adresa", "strasse", "straße", "adresse"),
    "psc": ("psc", "psč", "zip", "zip code", "postal code", "postcode", "plz"),
    "mesto": ("mesto", "město", "city", "town", "obec", "ort", "stadt"),
}


def klic_sloupce(hlavicka):
    h = bez_diakritiky(hlavicka or "").strip().lower()
    for klic, varianty in MAPOVANI_SLOUPCU.items():
        if h in {bez_diakritiky(v).lower() for v in varianty}:
            return klic
    return None


def nacti_vstup(cesta, sloupec_nazvu=None):
    pripona = os.path.splitext(cesta)[1].lower()
    radky = []

    if pripona in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Pro cteni XLSX nainstalujte openpyxl:  pip install openpyxl")
        ws = load_workbook(cesta, read_only=True, data_only=True).active
        it = ws.iter_rows(values_only=True)
        hlavicka = [str(c) if c is not None else "" for c in next(it, [])]
        mapa = {}
        for i, h in enumerate(hlavicka):
            mapa[i] = "nazev" if (sloupec_nazvu and h == sloupec_nazvu) else klic_sloupce(h)
        if "nazev" not in mapa.values():
            mapa[0] = "nazev"
        for r in it:
            zaznam = {}
            for i, hodnota in enumerate(r):
                k = mapa.get(i)
                if k and hodnota is not None:
                    zaznam[k] = str(hodnota).strip()
            if zaznam.get("nazev") or zaznam.get("ico") or zaznam.get("dic"):
                radky.append(zaznam)
        return radky

    with open(cesta, encoding="utf-8-sig", newline="") as f:
        vzorek = f.read(4096)
        f.seek(0)
        if pripona == ".txt" or not any(d in vzorek for d in ";,\t"):
            return [{"nazev": r.strip()} for r in f if r.strip()]
        try:
            dialekt = csv.Sniffer().sniff(vzorek, delimiters=";,\t|")
        except csv.Error:
            dialekt = csv.excel
            dialekt.delimiter = ";"
        for r in csv.DictReader(f, dialect=dialekt):
            zaznam = {}
            for h, hodnota in r.items():
                k = "nazev" if (sloupec_nazvu and h == sloupec_nazvu) else klic_sloupce(h)
                if k and hodnota:
                    zaznam[k] = str(hodnota).strip()
            if not zaznam and r:
                prvni = next(iter(r.values()), None)
                if prvni:
                    zaznam = {"nazev": str(prvni).strip()}
            if zaznam.get("nazev") or zaznam.get("ico") or zaznam.get("dic"):
                radky.append(zaznam)
    return radky


SLOUPCE_ZAKLAD = [
    ("jmeno", "Jméno"), ("ulice", "Ulice"), ("psc", "PSČ"), ("mesto", "Město"),
    ("zeme", "Země"), ("ico", "IČO"), ("dic", "DIČ"),
    ("nace", "NACE"), ("nace_popis", "NACE popis"),
    ("kod_kategorie", "Kód kategorie"), ("skupina", "Skupina"),
    ("kategorie", "Kategorie dodavatele"),
]
SLOUPCE_DOPLNKY = [
    ("zdroj_kategorie", "Zařazeno podle"), ("zdroj", "Zdroj dat"), ("shoda", "Shoda názvu"),
    ("stav", "Stav"), ("hledany_nazev", "Hledaný název"), ("region", "Region"),
    ("lei", "LEI"), ("reg_cislo", "Registrační číslo"), ("reg_rejstrik", "Rejstřík"),
    ("pravni_forma", "Právní forma"), ("datum_vzniku", "Datum vzniku"),
    ("dic_overeno", "DIČ ověřeno (VIES)"), ("nace_vse", "NACE (všechny)"),
    ("nace_zdroj", "NACE - zdroj"), ("nace_llm", "NACE (LLM)"),
    ("klasifikace", "Klasifikace (US NAICS)"),
    ("odkaz", "Odkaz na rejstřík"), ("poznamka", "Poznámka"),
]

# Rezim --jen-id: prvni sloupce jsou zamerne stejne jako vzorovy vstup
# (Nazev/ICO/DIC/Zeme), aby sel vystup rovnou pouzit jako vstup druheho,
# plneho behu - staci zkontrolovat/opravit doplnene ICO.
SLOUPCE_ID = [
    ("hledany_nazev", "Název"), ("identifikator", "IČO"), ("dic", "DIČ"), ("zeme", "Země"),
    ("jmeno", "Nalezené jméno"), ("reg_rejstrik", "Typ čísla / rejstřík"),
    ("shoda", "Shoda názvu"), ("stav", "Stav"), ("zdroj", "Zdroj dat"),
    ("poznamka", "Poznámka"),
]

SIRKY = {"Jméno": 40, "Ulice": 30, "PSČ": 9, "Město": 20, "Země": 7, "IČO": 12, "DIČ": 15,
         "NACE": 9, "NACE popis": 34, "Kód kategorie": 13, "Skupina": 24,
         "Kategorie dodavatele": 42, "Zařazeno podle": 14, "Zdroj dat": 12, "Shoda názvu": 11,
         "Stav": 12, "Hledaný název": 34, "Region": 18, "LEI": 22,
         "Registrační číslo": 18, "Rejstřík": 20, "Právní forma": 14,
         "NACE - zdroj": 22, "NACE (LLM)": 12, "Klasifikace (US NAICS)": 34,
         "Datum vzniku": 13, "DIČ ověřeno (VIES)": 16, "NACE (všechny)": 30,
         "Odkaz na rejstřík": 46, "Poznámka": 70,
         "Název": 34, "Nalezené jméno": 34, "Typ čísla / rejstřík": 20,
         "Shoda NACE (divize)": 20}


# ---------------------------------------------------------------------------
# Rucni zarazeni pres LLM chat (MS Copilot, ChatGPT...) - bez API klice
#
# Nezname obory (XXX-00) se daji dohledat pres LLM, ale bez programoveho
# pristupu k nemu (jen chatove okno) to nejde zapojit primo do behu skriptu.
# Misto toho nastroj pripravi soubor k rucnimu vlozeni do chatu a pak nacte
# odpoved zpet - stejny princip jako --jen-id, jen mezikrok dela clovek
# v chatu mimo skript.
# ---------------------------------------------------------------------------

def _potrebuje_llm_pomoc(z):
    """Firmy bez kategorie (XXX-00), nebo s jedinym zapsanym NACE, ktery je jen
    obecny/podpurny kod (viz PODPURNE_NACE a nace_nejisty) - u obou pripadu
    ma smysl dohledat obor rucne/LLM."""
    return (z.kod_kategorie == taxonomie.VYCHOZI_KOD or z.nace_nejisty) and (z.hledany_nazev or z.jmeno)


def zapis_export_llm(zaznamy, cesta, kategorie_ciselnik=None):
    """
    Vypise firmy bez spolehliveho oboru do textoveho souboru pripraveneho na
    vlozeni do LLM chatu. Narozdil od drivejsi verze se LLM neptame primo na
    nasi vlastni kategorii (vyzadovalo by to, aby LLM spravne pochopil nasi
    ~95kategorii taxonomii jen z jednoho vypisu v promptu), ale na standardni
    NACE kod - tu klasifikaci LLM uz dobre zna z trenovacich dat. Kategorii
    z NACE pak dopocita stejny overeny mechanismus (taxonomie.zarad), jaky
    se pouziva pro skutecny NACE z rejstriku - viz pouzij_nace_mapu().
    Vraci pocet vypsanych firem.
    """
    nevyresene = [z for z in zaznamy if _potrebuje_llm_pomoc(z)]
    if not nevyresene:
        return 0

    radky = [
        "U kazde z firem nize uved jeji skutecny hlavni obor podnikani jako "
        "NACE Rev. 2 kod (mezinarodni NACE, ceska varianta CZ-NACE nebo "
        "obdoba v jine zemi - staci uroven divize/skupiny, napr. 4791 nebo "
        "62.01). Nekterym firmam mame uz zapsany NACE, je to ale jen obecny/"
        "formalni udaj (napr. pronajem nemovitosti) - u nich over, jestli "
        "jde o provozni firmu s jinou skutecnou cinnosti, nebo skutecne "
        "o majetkovou/holdingovou entitu, kde je zapsany udaj spravny (v tom "
        "pripade napis zpet ten stejny kod). Pokud si u firmy nejsi jistý/á "
        "nebo o ni nic nenajdes, napis misto kodu 'neznamo' - chybejici "
        "udaj je lepsi nez neopodstatneny odhad.",
        "",
        "FIRMY K DOHLEDÁNÍ (%d):" % len(nevyresene),
    ]
    for z in nevyresene:
        udaje = [x for x in (z.zeme, z.ulice, z.psc, z.mesto) if x]
        kontext = " [uz zapsany NACE: %s %s]" % (z.nace, z.nace_popis or "") if z.nace_nejisty and z.nace else ""
        radky.append("- %s%s%s" % (z.hledany_nazev or z.jmeno,
                                   " (%s)" % ", ".join(udaje) if udaje else "", kontext))

    radky += [
        "",
        "Odpověz přesně v tomto formátu, jeden řádek na firmu, oddělovač ';', "
        "beze změny pořadí a bez dalšího textu okolo:",
        "Původní název;NACE kód;Stručné zdůvodnění",
    ]
    with open(cesta, "w", encoding="utf-8") as f:
        f.write("\n".join(radky))
    return len(nevyresene)


def nacti_nace_mapu(cesta):
    """Nacte rucne/LLM dohledany NACE (Nazev;NACE[;...]) z odpovedi LLM chatu."""
    mapa = {}
    with open(cesta, encoding="utf-8-sig", newline="") as f:
        vzorek = f.read(4096)
        f.seek(0)
        try:
            dialekt = csv.Sniffer().sniff(vzorek, delimiters=";,\t|")
        except csv.Error:
            dialekt = csv.excel
            dialekt.delimiter = ";"
        for radek in csv.reader(f, dialect=dialekt):
            if len(radek) < 2:
                continue
            nazev = radek[0].strip()
            kod = re.sub(r"\D", "", radek[1])
            if not nazev or not kod:
                continue
            klic = normalizuj_nazev(nazev)
            if klic:
                mapa[klic] = kod
    return mapa


def pouzij_nace_mapu(zaznamy, mapa, nace_mapa=None, kategorie_ciselnik=None, mapa_oboru=None):
    """
    Aplikuje rucne/LLM dohledany NACE na zaznamy, ktere ho potrebuji (viz
    _potrebuje_llm_pomoc) - kategorii z nej dopocita taxonomie.zarad(), stejne
    jako u skutecneho NACE z rejstriku, takze LLM neresi nasi vlastni
    taxonomii, jen (pro nej znamejsi) standardni NACE klasifikaci.
    Vraci pocet zmen.
    """
    zmeny = 0
    for z in zaznamy:
        if not _potrebuje_llm_pomoc(z):
            continue
        klic = normalizuj_nazev(z.hledany_nazev or z.jmeno)
        kod_nace = mapa.get(klic)
        if not kod_nace:
            continue
        z.nace_llm = kod_nace
        k = taxonomie.zarad(nace=kod_nace, mapa=nace_mapa, kategorie=kategorie_ciselnik,
                            mapa_oboru=mapa_oboru)
        if k["kod"] == taxonomie.VYCHOZI_KOD:
            continue
        z.kod_kategorie = k["kod"]
        z.kategorie = k["kategorie"]
        z.skupina = k["skupina"]
        z.zdroj_kategorie = "rucne (LLM pres NACE)"
        zmeny += 1
    return zmeny


def _stylizuj_hlavicku_xlsx(ws, hlavicka, sirky=None):
    """Spolecny vzhled hlavicky pro vsechny XLSX vystupy - tucne bile pismo na
    tmavem podkladu, zamrazeny prvni radek, autofilter, sirky sloupcu podle SIRKY."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    sirky = sirky if sirky is not None else SIRKY
    hlavicka_font = Font(bold=True, color="FFFFFF")
    vypln = PatternFill("solid", fgColor="2E4A62")
    for b in ws[1]:
        b.font = hlavicka_font
        b.fill = vypln
        b.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, n in enumerate(hlavicka, 1):
        ws.column_dimensions[get_column_letter(i)].width = sirky.get(n, 16)
    return hlavicka_font, vypln


def zapis_vystup(zaznamy, cesta, oddelovac=";", kompakt=False, jen_id=False):
    if jen_id:
        sloupce = SLOUPCE_ID
    else:
        sloupce = SLOUPCE_ZAKLAD + ([] if kompakt else SLOUPCE_DOPLNKY)
    hlavicka = [n for _, n in sloupce]
    radky = [[getattr(z, k, "") or "" for k, _ in sloupce] for z in zaznamy]

    if os.path.splitext(cesta)[1].lower() != ".xlsx":
        with open(cesta, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f, delimiter=oddelovac)
            w.writerow(hlavicka)
            w.writerows(radky)
        return

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("Pro zapis XLSX nainstalujte openpyxl:  pip install openpyxl\n"
                 "(nebo zvolte vystup s priponou .csv)")

    wb = Workbook()
    ws = wb.active
    ws.title = "Dodavatelé"
    ws.append(hlavicka)
    for r in radky:
        ws.append(r)
    hlavicka_font, vypln = _stylizuj_hlavicku_xlsx(ws, hlavicka)

    # druhy list: ciselnik pouzite taxonomie (nema smysl v rezimu --jen-id,
    # kde jde jen o dohledani identifikacniho cisla, ne o zarazeni)
    if not jen_id:
        ws2 = wb.create_sheet("Číselník kategorií")
        ws2.append(["Kód", "Skupina", "Kategorie dodavatele", "Počet dodavatelů"])
        pocty = {}
        for z in zaznamy:
            pocty[z.kod_kategorie] = pocty.get(z.kod_kategorie, 0) + 1
        for kod, skupina, nazev in taxonomie.prehled_kategorii():
            ws2.append([kod, skupina, nazev, pocty.get(kod, 0)])
        for b in ws2[1]:
            b.font = hlavicka_font
            b.fill = vypln
        ws2.freeze_panes = "A2"
        for i, s in enumerate((14, 28, 48, 16), 1):
            ws2.column_dimensions[get_column_letter(i)].width = s

        # treti list: ciselnik CZ-NACE divizi (2 cislice) - aby si klient mohl
        # dohledat, co ktery kod ve sloupci NACE znamena, bez hledani mimo sesit
        ws3 = wb.create_sheet("Číselník NACE")
        ws3.append(["Kód (divize)", "Název"])
        for kod, nazev in taxonomie.prehled_nace_divizi():
            ws3.append([kod, nazev])
        for b in ws3[1]:
            b.font = hlavicka_font
            b.fill = vypln
        ws3.freeze_panes = "A2"
        for i, s in enumerate((14, 60), 1):
            ws3.column_dimensions[get_column_letter(i)].width = s

    wb.save(cesta)


# ---------------------------------------------------------------------------
# Komparace NACE s externim zdrojem (napr. rucni/AI doplneni od kolegy) -
# porovna nas sloupec NACE se sloupcem, ktery nekdo dalsi pridal do jiz
# vygenerovaneho vystupu. Na rozdil od nacti_vstup() nic ze vstupu nezahazuje -
# potrebujeme zachovat i sloupce, o kterych nastroj sam o sobe nic nevi.
# ---------------------------------------------------------------------------

def nacti_tabulku(cesta):
    """Precte XLSX/CSV beze schematu - vrati (hlavicka, radky), vsechny sloupce
    tak, jak jsou, vc. tech, ktere nacti_vstup() by jako neznamy zahodil."""
    pripona = os.path.splitext(cesta)[1].lower()

    if pripona in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Pro cteni XLSX nainstalujte openpyxl:  pip install openpyxl")
        ws = load_workbook(cesta, read_only=True, data_only=True).active
        it = ws.iter_rows(values_only=True)
        hlavicka = [str(c) if c is not None else "" for c in next(it, [])]
        radky = [["" if c is None else str(c).strip() for c in r] for r in it]
        return hlavicka, radky

    with open(cesta, encoding="utf-8-sig", newline="") as f:
        vzorek = f.read(4096)
        f.seek(0)
        try:
            dialekt = csv.Sniffer().sniff(vzorek, delimiters=";,\t|")
        except csv.Error:
            dialekt = csv.excel
            dialekt.delimiter = ";"
        r = list(csv.reader(f, dialect=dialekt))
    if not r:
        return [], []
    return r[0], r[1:]


def zapis_tabulku(hlavicka, radky, cesta):
    """Zapise obecnou tabulku (hlavicka + radky) do XLSX/CSV se stejnym
    vzhledem hlavicky jako zapis_vystup()."""
    if os.path.splitext(cesta)[1].lower() != ".xlsx":
        with open(cesta, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(hlavicka)
            w.writerows(radky)
        return

    try:
        from openpyxl import Workbook
    except ImportError:
        sys.exit("Pro zapis XLSX nainstalujte openpyxl:  pip install openpyxl\n"
                 "(nebo zvolte vystup s priponou .csv)")
    wb = Workbook()
    ws = wb.active
    ws.title = "Komparace"
    ws.append(hlavicka)
    for r in radky:
        ws.append(r)
    _stylizuj_hlavicku_xlsx(ws, hlavicka)
    wb.save(cesta)


_NACE_KOD_RE = re.compile(r"\b\d+(?:\.\d+)?\b")


def _nace_divize_vsechny(text):
    """
    Vytahne vsechny NACE divize (prvni 2 cislice kazdeho kodu) z textu.
    Sloupec od kolegy/AI casto neni "holy" kod, ale text jako
    '26.11 Vyroba počítačů... + 46.52 Velkoobchod s počítači' - s teckovanou
    notaci i s vic kody v jedne bunce najednou (kdyz popis firmy pokryva vic
    cinnosti). Bere se mnozina vsech nalezenych divizi, ne jen prvni.
    """
    divize = set()
    for kod in _NACE_KOD_RE.findall(str(text or "")):
        cislice = kod.replace(".", "")
        if cislice:
            divize.add(cislice[:2])
    return divize


def porovnej_nace_divize(a, b):
    """
    Porovna dva NACE udaje na urovni divize (prvni 2 cislice) - odpousti
    rozdily v detailnejsi urovni mezi dvema ruznymi zdroji i teckovanou
    notaci/popisny text (viz _nace_divize_vsechny). Shoda = aspon jedna
    spolecna divize. Vraci True/False, nebo None, kdyz jedne ze stran
    zadny rozpoznatelny kod chybi (nelze vyhodnotit).
    """
    da = _nace_divize_vsechny(a)
    db = _nace_divize_vsechny(b)
    if not da or not db:
        return None
    return bool(da & db)


def _najdi_sloupec(hlavicka, nazev):
    try:
        return hlavicka.index(nazev)
    except ValueError:
        raise SystemExit(
            "sloupec '%s' ve vstupu nenalezen - existujici sloupce: %s"
            % (nazev, ", ".join(hlavicka)))


def nacti_stavy_k_obnove(cesta, spatne_stavy=(STAV_NENALEZENO, STAV_OVERIT, STAV_CHYBA)):
    """
    Precte drivejsi vystup teto aplikace a vrati mnozinu hodnot sloupce
    "Hledany nazev", jejichz Stav byl spatny (NENALEZENO/OVERIT/CHYBA) -
    pro --obnovit-nenalezene, aby se pri opakovanem behu nad stejnym
    vstupem vynutil cerstvy dotaz jen pro tyto konkretni firmy, misto mazani
    cele kese (ktera by zahodila i spravne nalezene zaznamy).
    """
    hlavicka, radky = nacti_tabulku(cesta)
    i_nazev = _najdi_sloupec(hlavicka, "Hledaný název")
    i_stav = _najdi_sloupec(hlavicka, "Stav")
    return {r[i_nazev] for r in radky if len(r) > i_stav and r[i_stav] in spatne_stavy}


def zpracuj_komparaci(cesta_vstup, sloupec_kolega, cesta_vystup, sloupec_nas="NACE"):
    """
    Porovna nas sloupec NACE se sloupcem, ktery do jiz vygenerovaneho vystupu
    pridal nekdo dalsi (napr. rucni/AI doplneni od kolegy) - radky se berou
    1:1 podle pozice, zadne parovani podle jmena/ICO (soubor uz je nas vlastni
    vystup jen s pridanym sloupcem navic). Shoda se pocita na urovni NACE
    divize (prvni 2 cislice), viz porovnej_nace_divize().
    """
    hlavicka, radky = nacti_tabulku(cesta_vstup)
    i_nas = _najdi_sloupec(hlavicka, sloupec_nas)
    i_kolega = _najdi_sloupec(hlavicka, sloupec_kolega)

    nova_hlavicka = hlavicka + ["Shoda NACE (divize)"]
    nove_radky = []
    celkem = shoda = nesoulad = chybi_nas = chybi_kolega = 0
    for r in radky:
        r = list(r) + [""] * (len(hlavicka) - len(r))    # kratsi radky (prazdne bunky na konci)
        celkem += 1
        kod_nas = r[i_nas] if i_nas < len(r) else ""
        kod_kolega = r[i_kolega] if i_kolega < len(r) else ""
        vysledek = porovnej_nace_divize(kod_nas, kod_kolega)
        if vysledek is None:
            znacka = ""
            if not _nace_divize_vsechny(kod_nas):
                chybi_nas += 1
            if not _nace_divize_vsechny(kod_kolega):
                chybi_kolega += 1
        elif vysledek:
            znacka = "ANO"
            shoda += 1
        else:
            znacka = "NE"
            nesoulad += 1
        nove_radky.append(r + [znacka])

    zapis_tabulku(nova_hlavicka, nove_radky, cesta_vystup)

    srovnatelnych = shoda + nesoulad
    print("Komparace NACE (uroven divize): %d radku celkem, %d srovnatelnych "
          "(obe strany maji kod)" % (celkem, srovnatelnych), file=sys.stderr)
    if srovnatelnych:
        print("  shoda: %d/%d (%.0f %%)" % (
            shoda, srovnatelnych, 100.0 * shoda / srovnatelnych), file=sys.stderr)
    if chybi_nas:
        print("  bez NACE u nas: %d" % chybi_nas, file=sys.stderr)
    if chybi_kolega:
        print("  bez NACE ve sloupci '%s': %d" % (sloupec_kolega, chybi_kolega), file=sys.stderr)
    print("Vysledek -> %s" % cesta_vystup, file=sys.stderr)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv=None):
    p = argparse.ArgumentParser(
        description="Obohaceni seznamu dodavatelu z verejnych rejstriku "
                    "(ARES, RPO SR, GLEIF, SEC EDGAR, Wikidata, VIES).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""priklady:
  python3 dodavatele.py dodavatele.csv -o vystup.xlsx
  python3 dodavatele.py seznam.txt -o vystup.csv --vies --workers 6
  python3 dodavatele.py --dump-taxonomy taxonomie.json     # export k uprave
  python3 dodavatele.py vstup.csv --taxonomy taxonomie.json
""")
    p.add_argument("vstup", nargs="?", help="CSV / XLSX / TXT se seznamem firem")
    p.add_argument("-o", "--vystup", default="dodavatele_vystup.xlsx",
                   help="vystupni soubor .xlsx nebo .csv (vychozi: %(default)s)")
    p.add_argument("--sloupec", help="nazev sloupce se jmenem firmy, pokud se neurci automaticky")
    p.add_argument("--oddelovac", default=";", help="oddelovac pro CSV vystup (vychozi: ;)")
    p.add_argument("--kompakt", action="store_true", help="jen zakladni sloupce")
    p.add_argument("--jen-id", action="store_true",
                   help="jen dohledat ICO/registracni cislo (bez plneho obohaceni) - "
                        "vystup jde rovnou pouzit jako vstup druheho, plneho behu")
    p.add_argument("--export-nezarazene", metavar="SOUBOR",
                   help="vypsat firmy bez spolehliveho oboru (bez kategorie, nebo jen "
                        "s obecnym/podpurnym NACE) do textu pripraveneho na vlozeni "
                        "do LLM chatu (Copilot, ChatGPT...) - ptame se na NACE kod, "
                        "ne primo na nasi kategorii")
    p.add_argument("--nace-mapa", metavar="SOUBOR",
                   help="CSV s rucne/LLM dohledanym NACE (Nazev;NACE) - odpoved z LLM "
                        "chatu, kategorie se z nej dopocita stejne jako u NACE "
                        "z rejstriku, aplikuje se pred zapisem vystupu")
    p.add_argument("--workers", type=int, default=4, help="pocet soubeznych dotazu (vychozi: 4)")
    p.add_argument("--prodleva", type=float, default=0.25,
                   help="min. prodleva mezi dotazy na jeden server v s (vychozi: 0.25)")
    p.add_argument("--pocet", type=int, default=30, help="kolik kandidatu nacist (vychozi: 30)")
    p.add_argument("--prah-ok", type=float, default=0.90,
                   help="skore shody nazvu pro automaticke prijeti (vychozi: 0.90)")
    p.add_argument("--prah-overit", type=float, default=0.72,
                   help="skore, pod kterym je zaznam nenalezeny (vychozi: 0.72)")
    p.add_argument("--vies", action="store_true", help="overit DIC v EU pres VIES (pomalejsi)")
    p.add_argument("--bez-ares", action="store_true")
    p.add_argument("--bez-sk", action="store_true")
    p.add_argument("--bez-fr", action="store_true")
    p.add_argument("--bez-sg", action="store_true")
    p.add_argument("--bez-tw", action="store_true")
    p.add_argument("--bez-de", action="store_true",
                   help="nepouzivat lokalni kopii nemeckeho Handelsregisteru")
    p.add_argument("--pripravit-de-rejstrik", action="store_true",
                   help="stahnout/rozbalit lokalni kopii nemeckeho Handelsregisteru "
                        "(%s) a skoncit" % os.path.basename(DE_REGISTER_DB))
    p.add_argument("--de-api-klic", default=os.environ.get("OPENREGISTER_API_KEY", ""),
                   help="API klic pro OpenRegister.de (openregister.de) - placena "
                        "sluzba se skutecnym oborem cinnosti (WZ2025) pro nemecke firmy; "
                        "ma prednost pred lokalnim Handelsregisterem. Klic se nikam "
                        "neuklada, jen se pouzije za behu - lze predat i pres "
                        "promennou prostredi OPENREGISTER_API_KEY")
    p.add_argument("--scoris-api-klic", default=os.environ.get("SCORIS_API_KEY", ""),
                   help="API klic pro Scoris (scoris.eu) - placena sluzba se skutecnym "
                        "NACE pro SE/FI/EE/LV/LT. Klic se nikam neuklada, jen se pouzije "
                        "za behu - lze predat i pres promennou prostredi SCORIS_API_KEY")
    p.add_argument("--bez-gb", action="store_true",
                   help="nepouzivat lokalni kopii Companies House (UK)")
    p.add_argument("--pripravit-gb-rejstrik", action="store_true",
                   help="stahnout/naimportovat lokalni kopii Companies House "
                        "(%s) a skoncit" % os.path.basename(GB_REGISTER_DB))
    p.add_argument("--bez-gleif", action="store_true")
    p.add_argument("--bez-gleif-popisy", action="store_true",
                   help="nepřekládat kódy GLEIF (rejstřík, právní forma) na text - rychlejší")
    p.add_argument("--bez-edgar", action="store_true")
    p.add_argument("--bez-wikidata", action="store_true")
    p.add_argument("--cache", default=".dodavatele_cache.json.gz",
                   help="soubor s kesi odpovedi (prazdny retezec = bez kese)")
    p.add_argument("--obnovit-nenalezene", metavar="SOUBOR",
                   help="drivejsi vystup (bez --kompakt) - firmy, ktere v nem mely "
                        "stav NENALEZENO/OVERIT/CHYBA, se pro tento beh vynucene "
                        "znovu dotazi (obejde kes jen pro ne), ostatni se beze "
                        "zmeny berou z kese")
    p.add_argument("--taxonomy", help="JSON soubor s vlastni taxonomii")
    p.add_argument("--dump-taxonomy", metavar="SOUBOR",
                   help="zapsat vestavenou taxonomii do JSON a skoncit")
    p.add_argument("--komparace", metavar="SOUBOR",
                   help="porovnat NACE se sloupcem, ktery nekdo pridal do jiz "
                        "vygenerovaneho vystupu (napr. rucni/AI doplneni od kolegy) - "
                        "vyzaduje --komparace-sloupec, jen porovna a skonci")
    p.add_argument("--komparace-sloupec", metavar="NAZEV",
                   help="nazev sloupce v --komparace souboru s porovnavanym NACE kodem")
    p.add_argument("--komparace-nas-sloupec", metavar="NAZEV", default="NACE",
                   help="nazev naseho sloupce s NACE kodem (vychozi: NACE)")
    p.add_argument("--komparace-vystup", metavar="SOUBOR",
                   help="kam zapsat vysledek komparace (vychozi: <--komparace>_komparace.<pripona>)")
    p.add_argument("--ua", default=UA, help="hlavicka User-Agent (SEC vyzaduje kontakt)")
    a = p.parse_args(argv)

    if a.dump_taxonomy:
        with open(a.dump_taxonomy, "w", encoding="utf-8") as f:
            json.dump(taxonomie.jako_json(), f, ensure_ascii=False, indent=2)
        print("Taxonomie zapsana do %s" % a.dump_taxonomy)
        return 0

    if a.komparace:
        if not a.komparace_sloupec:
            p.error("--komparace vyzaduje --komparace-sloupec")
        zaklad, pripona = os.path.splitext(a.komparace)
        vystup = a.komparace_vystup or ("%s_komparace%s" % (zaklad, pripona or ".xlsx"))
        zpracuj_komparaci(a.komparace, a.komparace_sloupec, vystup,
                          sloupec_nas=a.komparace_nas_sloupec)
        return 0

    if a.pripravit_de_rejstrik:
        de_pripravit_databazi()
        return 0

    if a.pripravit_gb_rejstrik:
        gb_pripravit_databazi()
        return 0

    if not a.vstup:
        p.error("chybi vstupni soubor (nebo pouzijte --dump-taxonomy)")

    try:
        spustit(a)
    except RuntimeError as e:
        sys.exit(str(e))
    return 0


def spustit(a, na_radek=None):
    """
    Provede plne obohaceni podle `a` (argparse.Namespace, nebo jakykoli
    objekt se stejnymi atributy jako CLI parametry main() - viz gui.py,
    ktery si vlastni Namespace staví rucne) a vrati seznam zpracovanych
    Zaznamu. Sdilene jadro pro CLI (main()) i pro desktopove GUI (gui.py),
    aby se logika behu nemusela udrzovat na dvou mistech.

    `na_radek(hotovo, celkem, zaznam)` se zavola po zpracovani kazdeho
    radku (navic k prubeznemu vypisu na stderr) - GUI si tim aktualizuje
    progress bar bez nutnosti parsovat konzolovy vystup.
    """
    mapa, ciselnik, mapa_oboru = None, None, None
    if a.taxonomy:
        with open(a.taxonomy, encoding="utf-8") as f:
            mapa, ciselnik, mapa_oboru = taxonomie.z_json(json.load(f))

    radky = nacti_vstup(a.vstup, a.sloupec)
    if not radky:
        raise RuntimeError("Ve vstupu %s nejsou zadne pouzitelne radky." % a.vstup)
    print("Nacteno %d radku z %s" % (len(radky), a.vstup), file=sys.stderr)

    klient = Klient(cache_soubor=a.cache or None, prodleva=a.prodleva, ua=a.ua)
    n = {"pocet": a.pocet, "prah_ok": a.prah_ok, "prah_overit": a.prah_overit,
         "vies": a.vies, "bez_ares": a.bez_ares, "bez_sk": a.bez_sk,
         "bez_fr": a.bez_fr, "bez_sg": a.bez_sg, "bez_tw": a.bez_tw,
         "bez_de": a.bez_de, "bez_gb": a.bez_gb, "openregister_klic": a.de_api_klic,
         "scoris_klic": a.scoris_api_klic,
         "bez_gleif": a.bez_gleif, "bez_gleif_popisy": a.bez_gleif_popisy,
         "bez_edgar": a.bez_edgar, "bez_wikidata": a.bez_wikidata,
         "mapa": mapa, "kategorie_ciselnik": ciselnik, "mapa_oboru": mapa_oboru}

    k_obnove = None
    if a.obnovit_nenalezene:
        k_obnove = nacti_stavy_k_obnove(a.obnovit_nenalezene)
        print("Vynucena obnova pro %d firem se spatnym stavem z '%s'" % (
            len(k_obnove), a.obnovit_nenalezene), file=sys.stderr)

    hotovo = [0]
    zamek = threading.Lock()

    def uloha(vstup):
        hledany = ((vstup.get("nazev") or "").strip()
                  or (vstup.get("ico") or "").strip()
                  or (vstup.get("dic") or "").strip())
        if k_obnove is not None and hledany in k_obnove:
            klient.zapni_obnovu()
            try:
                z = zpracuj_radek(vstup, klient, n)
            finally:
                klient.vypni_obnovu()
        else:
            z = zpracuj_radek(vstup, klient, n)
        with zamek:
            hotovo[0] += 1
            print("  [%d/%d] %-42.42s -> %-11s %s" % (
                hotovo[0], len(radky), z.hledany_nazev, z.stav,
                "%s %s" % (z.kod_kategorie, z.kategorie)), file=sys.stderr)
            if na_radek is not None:
                na_radek(hotovo[0], len(radky), z)
        return z

    with ThreadPoolExecutor(max_workers=max(1, a.workers)) as ex:
        zaznamy = list(ex.map(uloha, radky))

    klient.uloz_cache()

    if a.nace_mapa:
        mapa_llm = nacti_nace_mapu(a.nace_mapa)
        zmeny = pouzij_nace_mapu(zaznamy, mapa_llm, nace_mapa=mapa, kategorie_ciselnik=ciselnik,
                                 mapa_oboru=mapa_oboru)
        print("Rucni/LLM zarazeni z %s: pouzito %d/%d" % (a.nace_mapa, zmeny, len(mapa_llm)),
              file=sys.stderr)

    if a.export_nezarazene:
        pocet = zapis_export_llm(zaznamy, a.export_nezarazene, ciselnik)
        if pocet:
            print("Export pro LLM chat -> %s (%d firem bez kategorie)" % (
                a.export_nezarazene, pocet), file=sys.stderr)
        else:
            print("Vsechny firmy maji kategorii, export pro LLM chat se nevytvaril.",
                  file=sys.stderr)

    zapis_vystup(zaznamy, a.vystup, a.oddelovac, a.kompakt, jen_id=a.jen_id)

    souhrn = {}
    for z in zaznamy:
        souhrn[z.stav] = souhrn.get(z.stav, 0) + 1
    print("\nHotovo -> %s" % a.vystup, file=sys.stderr)
    print("Souhrn: " + ", ".join("%s=%d" % kv for kv in sorted(souhrn.items())), file=sys.stderr)
    nezarazeno = [z for z in zaznamy if z.kod_kategorie == taxonomie.VYCHOZI_KOD]
    if nezarazeno:
        print("Bez kategorie (%d): %s" % (
            len(nezarazeno), ", ".join(z.hledany_nazev for z in nezarazeno[:10])), file=sys.stderr)
    k_kontrole = [z for z in zaznamy if z.stav != STAV_OK]
    if k_kontrole:
        print("K rucni kontrole (%d): %s" % (
            len(k_kontrole), ", ".join(z.hledany_nazev for z in k_kontrole[:10])), file=sys.stderr)
    return zaznamy


if __name__ == "__main__":
    sys.exit(main())
